from django.shortcuts import render,redirect,get_object_or_404
import pandas as pd
import csv
import io
from .models import User, Item
from .forms import UploadForm, ReportsForm
from django.contrib import auth
from django.contrib.auth import login, authenticate
from django.contrib.auth import logout
# from django.contrib.auth.models import User
from django.urls import reverse_lazy
from django.utils.safestring import mark_safe
import sys
from time import strftime
import os
import pyodbc
import sqlalchemy as sa
from sqlalchemy import create_engine
from urllib.parse import quote_plus
import datetime
import locale
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from openpyxl.styles import Font,NamedStyle, Border, Side, Alignment
from .connect import connect
from .backend import update_db,update_school,update_fy
from openpyxl.drawing.image import Image
from django.contrib.auth.decorators import login_required
from . import modules
from .decorators import permission_required,custom_login_required
from config import settings
from django.contrib.auth.hashers import make_password,check_password
from django.contrib import messages
import csv
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient
from io import BytesIO
from . import new_views
from django.urls import reverse
import hashlib
from django.http import HttpResponseBadRequest


SCHOOLS = settings.SCHOOLS
db = settings.db
schoolCategory = settings.schoolCategory
schoolMonths = settings.schoolMonths

# Get the current date
current_date = datetime.now()
# Extract the month number from the current date
month_number = current_date.month
curr_year = current_date.year


def updatedb(request):
    if request.method == 'POST':
        update_db()
    return redirect('/dashboard/advantage')

def updatefy(request, school, year=""):
    if request.method == 'POST':
        if year:
            print("that")
            update_fy(school,year)
        else:
            print("this")
            year = curr_year
            update_fy(school,year)
    return redirect(f'/dashboard/{school}')


def updateschool(request,school):
    if request.method == 'POST':
        update_school(school)
    return redirect(f'/dashboard/{school}')

def loginView(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        cnxn = connect()
        cursor = cnxn.cursor()
        cursor.execute("SELECT * FROM [dbo].[User] WHERE Username = ?", (username,))
        user_row = cursor.fetchone()
        
        if user_row and check_password(password, user_row[1]):
                if user_row[2] == 'admin':
                    role = user_row[2]
                    request.session['user_role'] = role
                    request.session['username'] = user_row[0]
                    try:
                        user = User.objects.get(username=username)
                    except:
                        user = User.objects.create_user(username=username)
                    login(request, user)
                    
                    return redirect('/home/schools')
                else:
                    role = user_row[2]
                    request.session['user_role'] = role
                    request.session['username'] = user_row[0]

                    try:
                        user = User.objects.get(username=username)
                    except:
                        user = User.objects.create_user(username=username)
                    login(request, user)
                   
                    return redirect(f'/dashboard/{role}')

        return redirect('login')

    elif request.method == "GET":
        return render(request, 'login.html')

def logoutView(request):
    logout(request)
    return redirect('login')

def change_password(request,school):
    print(request)
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        username = request.POST.get('username')
      
        if password1 == password2:
      
            hashed_password = make_password(password2)
            cnxn = connect()
            cursor = cnxn.cursor()
            query = ("UPDATE [dbo].[User] SET Password = ? WHERE Username = ? ")
            cursor.execute(query,(hashed_password,username))
            cnxn.commit()
            messages.success(request, 'Password has been changed successfully.')
            return redirect(f'/dashboard/{school}')

        else:
            messages.error(request, 'Passwords do not match.')


    return redirect(f'/dashboard/{school}')


    
# def update_row(request,school,year):
#     if request.method == 'POST':
#         print(request)
#         print(school_year)
        
#         try:
#             cnxn = connect()
#             cursor = cnxn.cursor()
#             updatefyes = request.POST.getlist('updatefye[]')  
#             updateids = request.POST.getlist('updateID[]')             

#             updatedata_list = []

            
#             for updatefye,updateid in zip(updatefyes, updateids):
#                 if updatefye.strip() and updateid.strip() :
#                     updatefye = float(updatefye.replace("$", "").replace(",", "").replace("(", "-").replace(")", ""))
#                     updatedata_list.append({
                       
#                         'updatefye': updatefye,
#                         'updateid':updateid,
                        
                        
                      
                        
#                     })
#             for data in updatedata_list:
                
#                 updatefye= data['updatefye']
#                 updateid=data['updateid']
               
                
          

#                 try:
#                     query = "UPDATE [dbo].[BS_FYE] SET FYE = ? WHERE BS_id = ? and school = ? "
#                     cursor.execute(query, (updatefye, updateid,school))
#                     cnxn.commit()
                   
#                 except Exception as e:
#                     print(f"Error updating bs_id={updateid}: {str(e)}")
            
            
#             cursor.close()
#             cnxn.close()
#             anchor_year = school_year
#             print(anchor_year)
#             context = modules.balance_sheet(school,anchor_year)
#             role = request.session.get('user_role')
#             context["role"] = role
#             username = request.session.get('username')
#             context["username"] = username
#             return redirect(request, "temps/balance-sheet.html", context)

#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': str(e)})

#     return JsonResponse({'status': 'error', 'message': 'Invalid request method'}) 
    

# def insert_row(request):
#     if request.method == 'POST':
#         print(request.POST)
#         try:
#             #<<<------------------ LR INSERT FUNCTION ---------------------->>>
#             funds = request.POST.getlist('fund[]')
#             objs = request.POST.getlist('obj[]')
#             descriptions = request.POST.getlist('Description[]')
#             budgets = request.POST.getlist('budget[]')

#             data_list_LR = []
#             for fund, obj, description, budget in zip(funds, objs, descriptions, budgets):
#                 if fund.strip() and obj.strip() and budget.strip():
#                     data_list_LR.append({
#                         'fund': fund,
#                         'obj': obj,
#                         'description': description,
#                         'budget': budget,
#                         'category': 'Local Revenue',
#                     })

          
#             cnxn = connect()
#             cursor = cnxn.cursor()

          
#             for data in data_list_LR:
#                 fund = data['fund']
#                 obj = data['obj']
#                 description = data['description']
#                 budget = data['budget']
#                 category = data['category']

#                 query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
#                 cursor.execute(query, (budget, fund, obj, description, category))
#                 cnxn.commit()

            



#             #<--------------UPDATE FOR LOCAL REVENUE,SPR,FPR----------
#             updatefunds = request.POST.getlist('updatefund[]')  
#             updatevalues = request.POST.getlist('updatevalue[]')
#             updateobjs = request.POST.getlist('updateobj[]')
            
            

#             updatedata_list = []

#             for updatefund,updatevalue,updateobj in zip(updatefunds, updatevalues,updateobjs):
#                 if updatefund.strip() and updatevalue.strip() :
#                     updatedata_list.append({
#                         'updatefund': updatefund,
#                         'updateobj':updateobj,
                        
#                         'updatevalue': updatevalue,
                        
#                     })
#             for data in updatedata_list:
#                 updatefund= data['updatefund']
#                 updateobj=data['updateobj']
                
#                 updatevalue = data['updatevalue']

#                 try:
#                     query = "UPDATE [dbo].[AscenderData_Advantage_Definition_obj] SET budget = ? WHERE fund = ? and obj = ? "
#                     cursor.execute(query, (updatevalue, updatefund,updateobj))
#                     cnxn.commit()
#                     print(f"Rows affected for fund={updatefund}: {cursor.rowcount}")
#                 except Exception as e:
#                     print(f"Error updating fund={updatefund}: {str(e)}")


#             #<---------------------------------- INSERT FOR SPR ---->>>
#             fundsSPR = request.POST.getlist('fundSPR[]')
#             objsSPR = request.POST.getlist('objSPR[]')
#             descriptionsSPR = request.POST.getlist('DescriptionSPR[]')
#             budgetsSPR = request.POST.getlist('budgetSPR[]')

#             data_list_SPR = []
#             for fund, obj, description, budget in zip(fundsSPR, objsSPR, descriptionsSPR, budgetsSPR):
#                 if fund.strip() and obj.strip() and budget.strip():
#                     data_list_SPR.append({
#                         'fund': fund,
#                         'obj': obj,
#                         'description': description,
#                         'budget': budget,
#                         'category': 'State Program Revenue',
#                     })
          
#             for data in data_list_SPR:
#                 fund = data['fund']
#                 obj = data['obj']
#                 description = data['description']
#                 budget = data['budget']
#                 category = data['category']

#                 query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
#                 cursor.execute(query, (budget, fund, obj, description, category))
#                 cnxn.commit()



def update_row(request,school,year):
    if request.method == 'POST':
        print(request)
        print(year)
        

        cnxn = connect()
        cursor = cnxn.cursor()
        updatefyes = request.POST.getlist('updatefye[]')  
        updateids = request.POST.getlist('updateID[]') 
        
        
         
        
        
        
        
        updatedata_list = []
        
        for updatefye,updateid in zip(updatefyes, updateids):
            if updatefye.strip() and updateid.strip() :
                updatefye = float(updatefye.replace("$", "").replace(",", "").replace("(", "-").replace(")", ""))
                updatedata_list.append({
                   
                    'updatefye': updatefye,
                    'updateid':updateid,
                    
                    
                  
                    
                })
        for data in updatedata_list:
            
            updatefye= data['updatefye']
            updateid=data['updateid']
           
            
        
            try:
                query = "UPDATE [dbo].[BS_FYE] SET FYE = ? WHERE BS_id = ? and school = ? and year = ?"
                cursor.execute(query, (updatefye, updateid, school, year))
                cnxn.commit()
            except Exception as e:
                print(f"Error updating bs_id={updateid}: {str(e)}")
        
        
        cursor.close()
        cnxn.close()
        anchor_year = year
        context = modules.balance_sheet(school,anchor_year)
        role = request.session.get('user_role')
        context["role"] = role
        username = request.session.get('username')
        context["username"] = username
        return render(request, "temps/balance-sheet.html", context)


    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}) 

def insert_row(request):
    if request.method == 'POST':
        print(request.POST)
        try:
            #<<<------------------ LR INSERT FUNCTION ---------------------->>>
            funds = request.POST.getlist('fund[]')
            objs = request.POST.getlist('obj[]')
            descriptions = request.POST.getlist('Description[]')
            budgets = request.POST.getlist('budget[]')

            data_list_LR = []
            for fund, obj, description, budget in zip(funds, objs, descriptions, budgets):
                if fund.strip() and obj.strip() and budget.strip():
                    data_list_LR.append({
                        'fund': fund,
                        'obj': obj,
                        'description': description,
                        'budget': budget,
                        'category': 'Local Revenue',
                    })

          
            cnxn = connect()
            cursor = cnxn.cursor()

          
            for data in data_list_LR:
                fund = data['fund']
                obj = data['obj']
                description = data['description']
                budget = data['budget']
                category = data['category']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (budget, fund, obj, description, category))
                cnxn.commit()

            



            #<--------------UPDATE FOR LOCAL REVENUE,SPR,FPR----------
            updatefunds = request.POST.getlist('updatefund[]')  
            updatevalues = request.POST.getlist('updatevalue[]')
            updateobjs = request.POST.getlist('updateobj[]')
            
            

            updatedata_list = []

            for updatefund,updatevalue,updateobj in zip(updatefunds, updatevalues,updateobjs):
                if updatefund.strip() and updatevalue.strip() :
                    updatedata_list.append({
                        'updatefund': updatefund,
                        'updateobj':updateobj,
                        
                        'updatevalue': updatevalue,
                        
                    })
            for data in updatedata_list:
                updatefund= data['updatefund']
                updateobj=data['updateobj']
                
                updatevalue = data['updatevalue']

                try:
                    query = "UPDATE [dbo].[AscenderData_Advantage_Definition_obj] SET budget = ? WHERE fund = ? and obj = ? "
                    cursor.execute(query, (updatevalue, updatefund,updateobj))
                    cnxn.commit()
                    print(f"Rows affected for fund={updatefund}: {cursor.rowcount}")
                except Exception as e:
                    print(f"Error updating fund={updatefund}: {str(e)}")


            #<---------------------------------- INSERT FOR SPR ---->>>
            fundsSPR = request.POST.getlist('fundSPR[]')
            objsSPR = request.POST.getlist('objSPR[]')
            descriptionsSPR = request.POST.getlist('DescriptionSPR[]')
            budgetsSPR = request.POST.getlist('budgetSPR[]')

            data_list_SPR = []
            for fund, obj, description, budget in zip(fundsSPR, objsSPR, descriptionsSPR, budgetsSPR):
                if fund.strip() and obj.strip() and budget.strip():
                    data_list_SPR.append({
                        'fund': fund,
                        'obj': obj,
                        'description': description,
                        'budget': budget,
                        'category': 'State Program Revenue',
                    })
          
            for data in data_list_SPR:
                fund = data['fund']
                obj = data['obj']
                description = data['description']
                budget = data['budget']
                category = data['category']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (budget, fund, obj, description, category))
                cnxn.commit()
            fundsFPR = request.POST.getlist('fundFPR[]')
            objsFPR = request.POST.getlist('objFPR[]')
            descriptionsFPR = request.POST.getlist('DescriptionFPR[]')
            budgetsFPR = request.POST.getlist('budgetFPR[]')


            data_list_FPR = []
            for fund, obj, description, budget in zip(fundsFPR, objsFPR, descriptionsFPR, budgetsFPR):
                if fund.strip() and obj.strip() and budget.strip():
                    data_list_FPR.append({
                        'fund': fund,
                        'obj': obj,
                        'description': description,
                        'budget': budget,
                        'category': 'Federal Program Revenue',
                    })
          
            for data in data_list_FPR:
                fund = data['fund']
                obj = data['obj']
                description = data['description']
                budget = data['budget']
                category = data['category']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (budget, fund, obj, description, category))
                cnxn.commit()
            
            #<---------------------------------- INSERT FOR Func ---->>>
            newfuncs = request.POST.getlist('newfunc[]')
            newDescriptionfuncs = request.POST.getlist('newDescriptionfunc[]')
            newBudgetfuncs = request.POST.getlist('newBudgetfunc[]')
            

            data_list_func = []
            for func, description, budget in zip(newfuncs, newDescriptionfuncs,newBudgetfuncs):
                if func.strip() and budget.strip():
                    data_list_func.append({
                        'func': func,
                        'description': description,
                        'budget': budget,
                        
                    })
          
            for data in data_list_func:
                func = data['func']
                description = data['description']
                budget = data['budget']
                

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_func] (budget, func, Description) VALUES (?, ?, ?)"
                cursor.execute(query, (budget, func, description))
                cnxn.commit()
            
            updatefunds = request.POST.getlist('updatefund[]')  
            updatevalues = request.POST.getlist('updatevalue[]')
            updateobjs = request.POST.getlist('updateobj[]')
            

            #<------------------ update for func func ---------------
            updatefuncfuncs = request.POST.getlist('updatefuncfunc[]')  
            updatefuncbudgets = request.POST.getlist('updatefuncbudget[]')
            


            updatedata_list_func = []

            for updatefunc,updatebudget in zip(updatefuncfuncs, updatefuncbudgets):
                if updatefunc.strip() and updatebudget.strip() and updatebudget.strip() != " ":
                    updatedata_list_func.append({
                        'updatefunc': updatefunc,
                        'updatebudget':updatebudget,
                        
                        
                    })
            for data in updatedata_list_func:
                updatefunc= data['updatefunc']
                updatebudget=data['updatebudget']
                
                

                try:
                    query = "UPDATE [dbo].[AscenderData_Advantage_Definition_func] SET budget = ? WHERE func = ? "
                    cursor.execute(query, (updatebudget, updatefunc))
                    cnxn.commit()
                    print(f"Rows affected for fund={updatefunc}: {cursor.rowcount}")
                except Exception as e:
                    print(f"Error updating fund={updatefunc}: {str(e)}")

                


            cursor.close()
            cnxn.close()

            return redirect('pl_advantage')

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def delete(request, fund, obj):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_obj] WHERE fund = ? AND obj = ?"
        cursor.execute(query, (fund, obj))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_func(request, func):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_func] WHERE func = ?"
        cursor.execute(query, (func))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def delete_bs(request, description, subcategory):
    try:
        
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_Balancesheet] WHERE Description = ? and Subcategory = ?"
        cursor.execute(query, (description,subcategory))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_bsa(request, obj, Activity):
    try:
     
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_ActivityBS] WHERE obj = ? and Activity = ?"
        cursor.execute(query, (obj,Activity))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# def first_advantagechart(request):
#     return render(request,'dashboard/advantage/first_advantagechart.html')

# def pl_advantagechart(request):
#     return render(request,'dashboard/advantage/pl_advantagechart.html')

# def pl_cumberlandchart(request):
#     return render(request,'dashboard/cumberland/pl_cumberlandchart.html')



def delete(request, fund, obj):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_obj] WHERE fund = ? AND obj = ?"
        cursor.execute(query, (fund, obj))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_func(request, func):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_func] WHERE func = ?"
        cursor.execute(query, (func))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def delete_bs(request, description, subcategory):
    try:
        
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_Balancesheet] WHERE Description = ? and Subcategory = ?"
        cursor.execute(query, (description,subcategory))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_bsa(request, obj, Activity):
    try:
     
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_ActivityBS] WHERE obj = ? and Activity = ?"
        cursor.execute(query, (obj,Activity))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# def first_advantagechart(request):
#     return render(request,'dashboard/advantage/first_advantagechart.html')

# def pl_advantagechart(request):
#     return render(request,'dashboard/advantage/pl_advantagechart.html')

# def pl_cumberlandchart(request):
#     return render(request,'dashboard/cumberland/pl_cumberlandchart.html')

def viewgl(request,fund,obj,yr,school,year,url):
    null_values = [None, 'null', 'None']
    print(request)
    try:
        print(year)
        print(url)
        
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""

        
        cnxn = connect()
        cursor = cnxn.cursor()
        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        filters = settings.filters
            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])

        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query}  fund = ? and obj = ? and AcctPer = ? "
                cursor.execute(query, (fund,obj,yr))
                
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? and MONTH(Date) = ? "
                cursor.execute(query, (fund,obj,date_object.month))
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? and Month = ? "
                cursor.execute(query, (fund,obj,yr))
                
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? and MONTH(PostingDate) = ? "
                cursor.execute(query, (fund,obj,date_object.month))


        
        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }

                        gl_data.append(row_dict)

        if request.path.startswith('/viewgl/'):
            total_bal = sum(float(row['Real']) for row in gl_data)
        else:
            total_bal = sum(float(row['Amount']) for row in gl_data)
    
        total_bal = format_value(total_bal)


        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewgl_all(request, school, year, url, yr=""):
    null_values = [None, 'null', 'None']
    data = json.loads(request.body)
    # do something about the yr
    if not yr:
        yr = ['09', '10', '11']
    else:
        yr = [yr]

    try:
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""

        
        cnxn = connect()
        cursor = cnxn.cursor()
        # date_string = f"{year}-09-01T00:00:00.0000000"
        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")
        
        fund_obj_query = " OR ".join(["(fund = ? AND obj = ?)" for _ in range(len(data))])
 
        values = []
        for row in data:
            values.append(row['fund'])
            values.append(row['obj'])
        
        filters = settings.filters
        
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            if url == 'acc':

                values.extend(yr)
                date_query = " OR ".join("AcctPer = ?" for _ in range(len(yr)))
                query = f"""
                SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query})
                """
                cursor.execute(query, values)
                
            else:
                values.extend([yr.map(lambda x: int(x))])
                date_query = " OR ".join("MONTH(Date) = ?" for _ in range(len(yr)))
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query}) "
                cursor.execute(query, values)
        else:
            filter_query = ' AND '.join( f"{column} NOT IN {value}" for column, value in filters['skyward'].items())
            if url == 'acc':
                values.extend(yr)
                date_query = " OR ".join("Month = ?" for _ in range(len(yr)))
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query}) "
                cursor.execute(query, values)
                
            else:
                values.extend([yr.map(lambda x: int(x))])
                date_query = " OR ".join("MONTH(PostingDate) = ?" for _ in range(len(yr)))
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query}) "
                cursor.execute(query, values)
        
        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                    gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19],
                        }
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        gl_data.append(row_dict)

        total_bal = 0    
        expend_key = "Real"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        # for row in gl_data:
        #     expend_str = row[expend_key]
        #     try:
        #         expend_value = float(expend_str)
        #         total_bal += expend_value
        #     except ValueError:
        #         pass

        total_bal = sum(float(row[expend_key]) for row in gl_data)
    
        total_bal = format_value(total_bal)

        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})



def viewgl_cumberland(request,fund,obj,yr):
    
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()


        
        
        query = "SELECT * FROM [dbo].[AscenderData_Cumberland] WHERE fund = ? and obj = ? and AcctPer = ?"
        cursor.execute(query, (fund,obj,yr))
        
        rows = cursor.fetchall()
    
        gl_data=[]
    
    
        for row in rows:
            date_str=row[11]

          

            # real = float(row[14]) if row[14] else 0
            # if real == 0:
            #     realformat = ""
            # else:
            #     realformat = "{:,.0f}".format(abs(real)) if real >= 0 else "({:,.0f})".format(abs(real))

            
            row_dict = {
                'fund':row[0],
                'func':row[1],
                'obj':row[2],
                'sobj':row[3],
                'org':row[4],
                'fscl_yr':row[5],
                'pgm':row[6],
                'edSpan':row[7],
                'projDtl':row[8],
                'AcctDescr':row[9],
                'Number':row[10],
                'Date':date_str,
                'AcctPer':row[12],
                'Est':row[13],
                'Real':row[14],
                'Appr':row[15],
                'Encum':row[16],
                'Expend':row[17],
                'Bal':row[18],
                'WorkDescr':row[19],
                'Type':row[20],
                'Contr':row[21]
            }

            gl_data.append(row_dict)
        
        total_bal = sum(float(row['Real']) for row in gl_data)
        total_bal = "{:,.0f}".format(abs(total_bal)) if total_bal >= 0 else "({:,.0f})".format(abs(total_bal))
        
        
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def insert_bs_advantage(request):
    if request.method == 'POST':
        print(request.POST)
        try:
            #<<<------------------ CURRENT ASSETS INSERT FUNCTION ---------------------->>>
            Descriptions = request.POST.getlist('Description[]')
            fyes = request.POST.getlist('fye[]')
            Activitys = request.POST.getlist('Activity[]')
           

            data_list_current_assets = []
            for Description, fye , Activity in zip(Descriptions, fyes,Activitys):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_current_assets.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Current Assets',
                        'Category': 'Assets',
                        'Activity': Activity,
                    })

          
            cnxn = connect()
            cursor = cnxn.cursor()


            for data in data_list_current_assets:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()

            
            #<<<------------------ Capital  ASSETS INSERT FUNCTION ---------------------->>>
            Descriptions_Capital_Assets = request.POST.getlist('Description_Capital_Assets[]')
            fyes_Capital_Assets = request.POST.getlist('fye_Capital_Assets[]')
            Activitys_Capital_Assets = request.POST.getlist('Activity_Capital_Assets[]')
           

            data_list_capital_assets = []
            for Description, fye , Activity in zip(Descriptions_Capital_Assets, fyes_Capital_Assets,Activitys_Capital_Assets):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_capital_assets.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Capital Assets, Net',
                        'Category': 'Assets',
                        'Activity': Activity,
                    })

          
     

          
            for data in data_list_capital_assets:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()

             
            #<<<------------------ CURRENT LIABILITIES INSERT FUNCTION ---------------------->>>
            Descriptions_CLs = request.POST.getlist('Description_CL[]')
            fyes_CLs = request.POST.getlist('fye_CL[]')
            Activitys_CLs = request.POST.getlist('Activity_CL[]')
           

            data_list_CL = []
            for Description, fye , Activity in zip(Descriptions_CLs, fyes_CLs, Activitys_CLs):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_CL.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Current Liabilities',
                        'Category': 'Liabilities and Net Assets',
                        'Activity': Activity,
                    })

            for data in data_list_CL:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()


            #<<<------------------ LONG TERM DEBT INSERT FUNCTION ---------------------->>>
            Descriptions_LTDs = request.POST.getlist('Description_LTD[]')
            fyes_LTDs = request.POST.getlist('fye_LTD[]')
            Activitys_LTDs = request.POST.getlist('Activity_LTD[]')
           

            data_list_LTD = []
            for Description, fye , Activity in zip(Descriptions_LTDs, fyes_LTDs, Activitys_LTDs):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_LTD.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Long Term Debt',
                        'Category': 'Debt',
                        'Activity': Activity,
                    })

            for data in data_list_LTD:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()

            
            #<<<------------------ BALANCE SHEET ACTIVITY INSERT FUNCTION ---------------------->>>
            Description_BSAs = request.POST.getlist('Description_BSA[]')
            obj_BSAs = request.POST.getlist('obj_BSA[]')
            Activity_BSAs = request.POST.getlist('Activity_BSA[]')
           

            data_list_BSA = []
            for Description, obj , Activity in zip(Description_BSAs, obj_BSAs, Activity_BSAs):
                if Description.strip() and obj.strip() and Activity.strip():
                    data_list_BSA.append({
                        'Description': Description,
                        'obj': obj,
                        'Activity': Activity,
                    })

            for data in data_list_BSA:
                Description = data['Description']
                obj = data['obj']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_ActivityBS] (Description, obj, Activity) VALUES (?, ?, ?)"
                cursor.execute(query, (Description, obj, Activity))
                cnxn.commit()
            

            
            #<--------------UPDATE FOR LOCAL REVENUE,SPR,FPR----------
            updatesubs = request.POST.getlist('updatesub[]')  
            updatedescs = request.POST.getlist('updatedesc[]')
            updatefyes = request.POST.getlist('updatefye[]')
            
            

            updatedata_bs = []

            for updatesub,updatedesc,updatefye in zip(updatesubs, updatedescs,updatefyes):
                if updatesub.strip() and updatedesc.strip()  and updatefye.strip():
                    updatedata_bs.append({
                        'updatesub': updatesub,
                        'updatedesc':updatedesc,
                        
                        'updatefye': updatefye,
                        
                    })
            for data in updatedata_bs:
                updatesub= data['updatesub']
                updatedesc=data['updatedesc']
                
                updatefye = data['updatefye']

                try:
                    query = "UPDATE [dbo].[AscenderData_Advantage_Balancesheet] SET FYE = ? WHERE Description = ? and Subcategory = ? "
                    cursor.execute(query, (updatefye, updatedesc,updatesub))
                    cnxn.commit()
                    print(f"Rows affected for fund={updatefye}: {cursor.rowcount}")
                except Exception as e:
                    print(f"Error updating fund={updatefye}: {str(e)}")
            
            


            return redirect('bs_advantage')

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



def viewgl_activitybs(request,yr,school,year,url):
    null_values = [None, 'null', 'None']
    data = json.loads(request.body)
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()

        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        filters = settings.filters

        if data["obj"]:
            obj_query = "(" + " OR ".join("obj = ?" for _ in data["obj"]) +")"
        else: 
            obj_query = "obj = ?"
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND {obj_query} and AcctPer = ? ; "
            
        else:
            filter_query = ' AND '.join( f"{column} NOT IN {value}" for column, value in filters['skyward'].items())
            query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND {obj_query} and Month = ? ; "
        query_values = [obj for obj in data['obj']] if data['obj'] else ['']
        query_values.extend([yr])
        cursor.execute(query, query_values)
        rows = cursor.fetchall()
    
        glbs_data=[]
    


        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    glbs_data.append(row_dict)

        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]
                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": float(row[19]) if row[19] not in null_values else '',
                        }
                        print(amount)
                        glbs_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        print(amount)
                        glbs_data.append(row_dict)
        print("hm")
        total_expend = 0 
        bal_key = "Bal"
        if school in schoolCategory["skyward"]:
            bal_key = "Amount"
        for row in glbs_data:
            expend_str = row[bal_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            
        
        
        print("hms")
        # total_bal = sum(row['Bal'] for row in glbs_data)
        total_bal = format_value(total_expend)
        
        context = { 
            'glbs_data':glbs_data,
            'total_bal':total_bal,

            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})



def viewglfunc(request,func,yr,school,year,url):
    null_values = [None, 'null', 'None']
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""
     
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()

        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")


        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and AcctPer = ? and obj != '6449' and Number != 'BEGBAL'; "
                cursor.execute(query, (func,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and MONTH(Date) = ? and obj != '6449' and Number != 'BEGBAL'; "
  
                cursor.execute(query, (func,date_object.month))
                
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and Month = ? and obj != '6449'; "
                cursor.execute(query, (func,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and obj != '6449' and MONTH(PostingDate) = ? "
                cursor.execute(query, (func,date_object.month))

        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:
                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }

                        # row_dict = {
                        #     "fund": row[0],
                        #     "func": row[2],
                        #     "obj": row[3],
                        #     "sobj": row[4],
                        #     "org": row[5],
                        #     "fscl_yr": row[6],
                        #     "Date": date,
                        #     "AcctPer": row[10],
                        #     "Amount": amount,
                        #     "Budget":row[20],
                        # }
                        gl_data.append(row_dict)

       
        total_expend = 0    
        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            

        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        
       
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglfunc_all(request,school,year,url, yr=""):
    null_values = [None, 'None', 'null']
    if not yr:
        yr = ['09', '10', '11']
    else:
        yr = [yr]
    data = json.loads(request.body)
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""
     
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()

        # date_string = f"{year}-09-01T00:00:00.0000000"
        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")
        query_func = " OR ".join(["func = ?" for _ in data])


        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            if url == 'acc':
                data.extend(yr)
                query_date = " OR ".join(["AcctPer = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and ({query_date}) and obj != '6449' and Number != 'BEGBAL'; "
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join(["MONTH(Date) = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and ({query_date}) and obj != '6449' and Number != 'BEGBAL'; "
  
                cursor.execute(query, data)
                
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            if url == 'acc':
                query_date = " OR ".join(["Month = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and ({query_date}) and obj != '6449'; "
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join(["MONTH(PostingDate) = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and obj != '6449' and ({query_date}) "
                cursor.execute(query, data)

        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        gl_data.append(row_dict)

       
        total_expend = 0    
        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            

        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        
       
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }
        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewgldna(request,func,yr,school,year,url):
    print(request)
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""
        
        year = int(year)
        FY_year_1 = year
        FY_year_2 = year + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()
        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")
        if school in schoolCategory["ascender"]:
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and AcctPer = ? and obj = '6449' and Number != 'BEGBAL'; "
                cursor.execute(query, (func,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and MONTH(Date) = ? and obj = '6449' and Number != 'BEGBAL'; "
  
                cursor.execute(query, (func,date_object.month))
                
        else:
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and Month = ? and obj = '6449'; "
                cursor.execute(query, (func,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and obj = '6449' and MONTH(PostingDate) = ? "
                cursor.execute(query, (func,date_object.month))

        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()
                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Date": date,
                            "AcctPer":acct_per_month,
                            "Amount": amount,
                            "Budget":row[20],
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Date": date,
                            "AcctPer": row[10],
                            "Amount": amount,
                            "Budget":row[20],
                        }
                        print(amount)
                        gl_data.append(row_dict)

       
        total_expend = 0    
        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            

        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        
       
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglexpense(request,obj,yr,school,year,url):
    null_values = [None, 'None', 'null']
    print(request)
    try:
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""
                
        cnxn = connect()
        cursor = cnxn.cursor()

        cursor.execute(f"SELECT  * FROM [dbo].{db[school]['code']};")
        rows = cursor.fetchall()

        data_expensebyobject=[]


        for row in rows:

            row_dict = {
                'obj':row[0],
                'Description':row[1],
                'budget':row[2],

                }

            data_expensebyobject.append(row_dict)

        cursor.execute(f"SELECT  * FROM [dbo].{db[school]['activities']};")
        rows = cursor.fetchall()

        data_activities=[]


        for row in rows:

        
            row_dict = {
                'obj':row[0],
                'Description':row[1],
                'Category':row[2],

                }

            data_activities.append(row_dict)
        


            
        
 
 
        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        # filters must be a dictionary with key = column and value = (tuple of values to filter)
        # this filter only works for categorical or string values

        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])

        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and AcctPer = ? "    
                cursor.execute(query, (obj,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and MONTH(Date) = ? "    
                cursor.execute(query, (obj,date_object.month))
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and Month = ? "    
                cursor.execute(query, (obj,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and MONTH(PostingDate) = ? "    
                cursor.execute(query, (obj,date_object.month))
        rows = cursor.fetchall()
    
        gl_data=[]
    
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0],
                            "T": row[1],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "PI": row[7],
                            "LOC": row[8],
                            "Date": date,
                            "AcctPer": row[10],
                            "Source": row[11],
                            "Subsource": row[12],
                            "Batch": row[13],
                            "Vendor": row[14],
                            "TransactionDescr": row[15],
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17],
                            "CheckDate": row[18],
                            "Amount": amount
                        }
                        gl_data.append(row_dict)


        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        total_expend = 0 
        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            
        
        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        

        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglexpense_all(request,school,year,url,yr=""):
    null_values = ['null', None, 'None']
    data = json.loads(request.body)
    if not yr:
        yr = ['09', '10', '11']
    else:
        yr = [yr]
    try:
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return ""
                
        cnxn = connect()
        cursor = cnxn.cursor()

        cursor.execute(f"SELECT  * FROM [dbo].{db[school]['code']};")
        rows = cursor.fetchall()

        data_expensebyobject=[]


        for row in rows:

            row_dict = {
                'obj':row[0],
                'Description':row[1],
                'budget':row[2],

                }

            data_expensebyobject.append(row_dict)

        cursor.execute(f"SELECT  * FROM [dbo].{db[school]['activities']};")
        rows = cursor.fetchall()

        data_activities=[]


        for row in rows:

        
            row_dict = {
                'obj':row[0],
                'Description':row[1],
                'Category':row[2],

                }

            data_activities.append(row_dict)
 
        # date_string = f"{year}-09-01T00:00:00.0000000"
        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        query_obj = " OR ".join("obj = ?" for _ in data)


        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])

        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                data.extend(yr)
                query_date = " OR ".join("AcctPer = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date}) "    
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join("MONTH(Date) = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date})"    
                cursor.execute(query, data)
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                data.extend(yr)
                query_date = " OR ".join("Month = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date}) "    
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join("MONTH(PostingDate) = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date}) "    
                cursor.execute(query, data)
        rows = cursor.fetchall()
    
        gl_data=[]
    
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                if invoice_date not in null_values:
                    invoice_date = invoice_date if invoice_date.year > 2021 else ''
                else:
                    invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer": row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        print(amount)
                        gl_data.append(row_dict)


        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        total_expend = 0 
        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            
        
        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        

        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def general_ledger_excel(request, school, start="", end=""):
    data = modules.general_ledger(school, start, end)['data3']

    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define your headers and add them to the worksheet
    headers = data[0].keys()
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Populate the worksheet with data
    for row_num, item in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=row_num, column=col_num).value = item[header]

    # Set the filename and mime type
    filename = os.path.join(settings.BASE_DIR,'finance','static', 'general_ledger.xlsx')

    # Save the workbook
    workbook.save(filename)

    # convert it to csv
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="download.csv"'

    writer = csv.writer(response)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)


    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = f'attachment; filename="download.csv"'

    # # Save the workbook
    # workbook.save(filename)

    return response


def generate_excel(request,school,anchor_year):
    cnxn = connect()
    cursor = cnxn.cursor()

    
    if anchor_year != curr_year :
        JSON_DIR = os.path.join(settings.BASE_DIR, "finance","json", str(anchor_year),  "excel", school)
    else:
        JSON_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "excel", school)
    
    with open(os.path.join(JSON_DIR, "data.json"), "r") as f:
        data = json.load(f)
    


    with open(os.path.join(JSON_DIR, "data2.json"), "r") as f:
        data2 = json.load(f)
    
    with open(os.path.join(JSON_DIR, "data3.json"), "r") as f:
        data3 = json.load(f)
    
    with open(os.path.join(JSON_DIR, "data_expensebyobject.json"), "r") as f:
        data_expensebyobject = json.load(f)
    
    with open(os.path.join(JSON_DIR, "data_activities.json"), "r") as f:
        data_activities = json.load(f)

    with open(os.path.join(JSON_DIR, "data_balancesheet.json"), "r") as f:
        data_balancesheet = json.load(f)
  
    with open(os.path.join(JSON_DIR, "data_activitybs.json"), "r") as f:
        data_activitybs = json.load(f)
  
    with open(os.path.join(JSON_DIR, "data_cashflow.json"), "r") as f:
        data_cashflow = json.load(f)

    with open(os.path.join(JSON_DIR, "data_charterfirst.json"), "r") as f:
        data_charterfirst = json.load(f)
 
    with open(os.path.join(JSON_DIR, "months.json"), "r") as f:
        months = json.load(f)
    with open(os.path.join(JSON_DIR, "totals.json"), "r") as f: #FOR PL
        totals = json.load(f)
    with open(os.path.join(JSON_DIR, "total_bs.json"), "r") as f: #FOR BS
        total_bs = json.load(f)

    school_name = SCHOOLS[school]

    template_path = os.path.join(settings.BASE_DIR, 'finance', 'static', 'template.xlsx')


    generated_excel_path = os.path.join(settings.BASE_DIR, 'finance', 'static', 'GeneratedExcel.xlsx')
    image_path = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','G.png' )

    img_g = Image(image_path)
    img_g.width = 50
    img_g.height = 50


    shutil.copyfile(template_path, generated_excel_path)

 
    workbook = openpyxl.load_workbook(generated_excel_path)
    sheet_names = workbook.sheetnames
    fontbold = Font(bold=True)


    first = sheet_names[0]
    pl = sheet_names[1]
    bs = sheet_names[2]
    cashflow= sheet_names[3]
    

    first_sheet = workbook[first]
    pl_sheet = workbook[pl]
    bs_sheet = workbook[bs]
    cashflow_sheet = workbook[cashflow]

    #------ FIRST DESIGN
    first_sheet.column_dimensions['A'].width = 46
    first_sheet.column_dimensions['B'].width = 20
    first_sheet.column_dimensions['C'].width = 9
    first_sheet.column_dimensions['D'].width = 20
    first_sheet.column_dimensions['E'].width = 38
    for row in range(3,26):
        first_sheet.row_dimensions[row].height = 30
    first_sheet.row_dimensions[1].height = 21
    first_sheet.row_dimensions[1].height = 21
    

    # image_cell = NamedStyle(name="image_cell", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    # normal_cell_bottom_border.border =Border(bottom=Side(border_style='thin'))
    # normal_cell_bottom_border.alignment = Alignment(horizontal='right', vertical='bottom')
    # normal_font_bottom_border = Font(name='Calibri', size=11, bold=False)
    # normal_cell_bottom_border.font = normal_font_bottom_border

    image_path1 = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','ofconcern.png' )
    image_path2 = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','ontrack.png' )
    image_path3 = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','atrisk.png' )


    image_list_concern = []
    image_list_track = []
    image_list_risk = []

    for i in range(1, 20):
  
        img_concern = Image(image_path1)
        img_track = Image(image_path2)
        img_risk = Image(image_path3)   

        image_list_concern.append(img_concern)
        image_list_track.append(img_track)  
        image_list_risk.append(img_risk) 

    
    start = 1
    first_start_row = 4
    for row in data_charterfirst:
        if row['school'] == school:
  # Create a new Image object
    
            first_sheet[f'A{start}'] = school_name
            start += 1
            first_sheet[f'A{start}'] = f'FY{months["FY_year_1"]}-{months["FY_year_2"]} Charter FIRST Forecasts of {months["last_month"]}'



    
    
            # Set the image position within the cell   
            first_sheet[f'B{first_start_row}'] = row['net_income_ytd']

            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['indicators']
            if row['indicators'].upper() == 'PASS':
                first_sheet.add_image(image_list_track[0],f'D{first_start_row}')

            else:
                first_sheet.add_image(image_list_risk[0],f'D{first_start_row}')


            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['net_assets']
            if row['net_assets'].upper() == 'PROJECTED':
    
                first_sheet.add_image(image_list_track[1],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[1],f'D{first_start_row}')
    

            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['days_coh']
            if row['days_coh'] > 60:
                first_sheet.add_image(image_list_track[2],f'D{first_start_row}')
            elif row['days_coh'] < 20:
                first_sheet.add_image(image_list_risk[2],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_concern[2],f'D{first_start_row}')
                


                
         
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['current_assets']
            if row['current_assets'] >= 2:
                first_sheet.add_image(image_list_track[3],f'D{first_start_row}')
            elif row['current_assets'] <= 1 :
                first_sheet.add_image(image_list_risk[3],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_concern[3],f'D{first_start_row}')

            
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['net_earnings']
            #first_sheet.add_image(image_list_track[17],f'D{first_start_row}') comment as of now
            
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['budget_vs_revenue']
            if row['budget_vs_revenue'].upper() == 'PROJECTED':   
                first_sheet.add_image(image_list_track[4],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[4],f'D{first_start_row}')
            

            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['total_assets']
            first_sheet.add_image(image_list_track[5],f'D{first_start_row}')
            
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['debt_service']
            first_sheet.add_image(image_list_risk[6],f'D{first_start_row}') 
            
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['debt_capitalization'] / 100
            first_sheet.add_image(image_list_track[7],f'D{first_start_row}')
           
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['ratio_administrative']
            first_sheet.add_image(image_list_track[8],f'D{first_start_row}')
          
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['ratio_student_teacher']
            if row['ratio_student_teacher'].lower() == 'not measured by dss':
                first_sheet.add_image(image_list_track[9],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[9],f'D{first_start_row}')
                
      
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['estimated_actual_ada']
            if row['estimated_actual_ada'].upper() == 'PROJECTED':   
                first_sheet.add_image(image_list_track[10],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[10],f'D{first_start_row}')
      
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['reporting_peims']
            if row['reporting_peims'].upper() == 'PROJECTED':   
                first_sheet.add_image(image_list_track[11],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[11],f'D{first_start_row}')
      
      
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['annual_audit']
            if row['annual_audit'].upper() == 'PROJECTED':   
                first_sheet.add_image(image_list_track[12],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[12],f'D{first_start_row}')

            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['post_financial_info']
            if row['post_financial_info'].upper() == 'PROJECTED':   
                first_sheet.add_image(image_list_track[13],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[13],f'D{first_start_row}')

            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['approved_geo_boundaries']
            if row['approved_geo_boundaries'].lower() == 'not measured by dss':
                first_sheet.add_image(image_list_track[14],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[14],f'D{first_start_row}')


            first_start_row += 1         
            first_sheet[f'B{first_start_row}'] = row['estimated_first_rating']
            if row['estimated_first_rating'] < 69:
                first_sheet.add_image(image_list_risk[15],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'F - Fail'
            elif row['estimated_first_rating'] < 80:
                first_sheet.add_image(image_list_concern[15],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'C - Meets Standard'
            elif row['estimated_first_rating'] < 90:
 
                first_sheet.add_image(image_list_track[16],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'B - Above Standard'
            else:
                first_sheet.add_image(image_list_track[15],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'A - Superior'

            
    

    #------- PL DESIGN


    pl_sheet.row_dimensions[1].height = 64
    pl_sheet.row_dimensions[3].height = 40
    for row in range(4,181):
        pl_sheet.row_dimensions[row].height = 19

    pl_sheet.column_dimensions['A'].width = 8
    pl_sheet.column_dimensions['B'].width = 28
    pl_sheet.column_dimensions['C'].width = 10
    pl_sheet.column_dimensions['D'].width = 16
    pl_sheet.column_dimensions['E'].width = 16
    pl_sheet.column_dimensions['F'].hidden = True
    pl_sheet.column_dimensions['G'].width = 16
    pl_sheet.column_dimensions['H'].width = 16
    pl_sheet.column_dimensions['I'].width = 16
    pl_sheet.column_dimensions['J'].width = 16
    pl_sheet.column_dimensions['K'].width = 16
    pl_sheet.column_dimensions['L'].width = 16
    pl_sheet.column_dimensions['M'].width = 16
    pl_sheet.column_dimensions['N'].width = 16
    pl_sheet.column_dimensions['O'].width = 16
    pl_sheet.column_dimensions['P'].width = 16
    pl_sheet.column_dimensions['Q'].width = 16
    pl_sheet.column_dimensions['R'].width = 16
    pl_sheet.column_dimensions['S'].width = 3
    pl_sheet.column_dimensions['T'].width = 17
    pl_sheet.column_dimensions['U'].width = 17
    pl_sheet.column_dimensions['V'].width = 12

    thin_border = Border(top=Side(border_style='thin'))
    currency_style = NamedStyle(name="currency_style", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style.border =Border(top=Side(border_style='thin'))
    currency_style.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style.font = currency_font


    normal_cell_bottom_border = NamedStyle(name="normal_cell_bottom_border", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell_bottom_border.border =Border(bottom=Side(border_style='thin'))
    normal_cell_bottom_border.alignment = Alignment(horizontal='right', vertical='bottom')
    normal_font_bottom_border = Font(name='Calibri', size=11, bold=False)
    normal_cell_bottom_border.font = normal_font_bottom_border

    normal_cell = NamedStyle(name="normal_cell", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell.alignment = Alignment(horizontal='right', vertical='bottom')
    normal_font = Font(name='Calibri', size=11, bold=False)
    normal_cell.font = normal_font


    currency_style_noborder = NamedStyle(name="currency_style_noborder", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style_noborder.alignment = Alignment(horizontal='right', vertical='top')
    currency_font_noborder = Font(name='Calibri', size=11, bold=True)
    currency_style_noborder.font = currency_font_noborder




    if school in schoolMonths["septemberSchool"]:


        for col in range(7, 20 ):
            col_letter = get_column_letter(col)
            pl_sheet.column_dimensions[col_letter].outline_level = 1
            pl_sheet.column_dimensions[col_letter].hidden = True

        last_number = months["last_month_number"]
        # PL START OF DESIGN
        if last_number <= 8:
            last_number += 11
        else:
            last_number -= 1

        for col in range(last_number,19):
            col_letter = get_column_letter(col)

      
            pl_sheet.column_dimensions[col_letter].outline_level = 2
            pl_sheet.column_dimensions[col_letter].hidden = True
           

        

        start_pl = 1
        pl_sheet[f'B{start_pl}'] = f'{school_name}\nFY{months["FY_year_1"]}-FY{months["FY_year_2"]} Statement of\nActivities as of {months["last_month"]}'
        start_pl += 2
        pl_sheet[f'E{start_pl}'] = f'{months["format_ytd_budget"]}% YTD \nBUDGET'
        pl_sheet[f'V{start_pl}'] = f'Var. {months["format_ytd_budget"]}'
        start_row = 5
        lr_row_end = None
        lr_row_start = start_row
        for row_data in data:
            if row_data['category'] == 'Local Revenue':
                all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'A{start_row}'] = row_data['fund']
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
    
                    pl_sheet[f'G{start_row}'] = -(row_data['total_real9'])
                    pl_sheet[f'H{start_row}'] = -(row_data['total_real10'])
                    pl_sheet[f'I{start_row}'] = -(row_data['total_real11'])
                    pl_sheet[f'J{start_row}'] = -(row_data['total_real12'])
                    pl_sheet[f'K{start_row}'] = -(row_data['total_real1'])
                    pl_sheet[f'L{start_row}'] = -(row_data['total_real2'])
                    pl_sheet[f'M{start_row}'] = -(row_data['total_real3'])
                    pl_sheet[f'N{start_row}'] = -(row_data['total_real4'])
                    pl_sheet[f'O{start_row}'] = -(row_data['total_real5'])
                    pl_sheet[f'P{start_row}'] = -(row_data['total_real6'])
                    pl_sheet[f'Q{start_row}'] = -(row_data['total_real7'])
                    pl_sheet[f'R{start_row}'] = -(row_data['total_real8'])
                    pl_sheet[f'T{start_row}'] = -(row_data['ytd_total'])
                    lr_row_end = start_row
                    start_row += 1

        if lr_row_end is not None:
            for col in range(4, 22):
                try:

                    cell = pl_sheet.cell(row=lr_row_end, column=col)

                    cell.style = normal_cell_bottom_border
                except KeyError as e:
                    print(f"Error hiding row {col}: {e}") 

            for row in range(lr_row_start, lr_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True

                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")           
        lr_end = start_row
        #local revenue total
        for col in range(2, 22):
            try:  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)

            cell.style = currency_style_noborder
        pl_sheet[f'B{start_row}'] = 'Local Revenue'
        pl_sheet[f'D{start_row}'] =  totals["total_ammended_lr"]
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_lr"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_lr"]["09"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_lr"]["10"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_lr"]["11"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_lr"]["12"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_lr"]["01"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_lr"]["02"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_lr"]["03"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_lr"]["04"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_lr"]["05"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_lr"]["06"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_lr"]["07"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_lr"]["08"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_lr"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue_lr"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'


        start_row += 1  



        spr_row_start = start_row
        spr_row_end = None
        for row_data in data:
            if row_data['category'] == 'State Program Revenue':
                all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell

                    pl_sheet[f'A{start_row}'] = row_data['fund']
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 

                    pl_sheet[f'G{start_row}'] = -(row_data['total_real9'])
                    pl_sheet[f'H{start_row}'] = -(row_data['total_real10'])
                    pl_sheet[f'I{start_row}'] = -(row_data['total_real11'])
                    pl_sheet[f'J{start_row}'] = -(row_data['total_real12'])
                    pl_sheet[f'K{start_row}'] = -(row_data['total_real1'])
                    pl_sheet[f'L{start_row}'] = -(row_data['total_real2'])
                    pl_sheet[f'M{start_row}'] = -(row_data['total_real3'])
                    pl_sheet[f'N{start_row}'] = -(row_data['total_real4'])
                    pl_sheet[f'O{start_row}'] = -(row_data['total_real5'])
                    pl_sheet[f'P{start_row}'] = -(row_data['total_real6'])
                    pl_sheet[f'Q{start_row}'] = -(row_data['total_real7'])
                    pl_sheet[f'R{start_row}'] = -(row_data['total_real8'])
                    pl_sheet[f'T{start_row}'] = -(row_data['ytd_total'])

                    pl_sheet[f'U{start_row}']= (row_data['variances'])
                    spr_row_end = start_row



                    start_row += 1


        if spr_row_end is not None:
            for col in range(4, 22):  # Columns G to U
                cell = pl_sheet.cell(row=spr_row_end, column=col)

                cell.style = normal_cell_bottom_border

            for row in range(spr_row_start, spr_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True

                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")    


        spr_end = start_row
        # STATE PROGRAM TOTAL      
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)        
            cell.style = currency_style_noborder
        pl_sheet[f'B{start_row}'] = 'State Program Revenue'
        pl_sheet[f'D{start_row}'] =  totals["total_ammended_spr"]    
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_spr"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_spr"]["09"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_spr"]["10"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_spr"]["11"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_spr"]["12"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_spr"]["01"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_spr"]["02"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_spr"]["03"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_spr"]["04"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_spr"]["05"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_spr"]["06"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_spr"]["07"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_spr"]["08"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_spr"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue_spr"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'


        start_row += 1

        fpr_row_end = None
        fpr_row_start = start_row
        for row_data in data:
            if row_data['category'] == 'Federal Program Revenue':
                all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 

                    pl_sheet[f'A{start_row}'] = row_data['fund']
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']             
                    pl_sheet[f'G{start_row}'] = -(row_data['total_real9'])
                    pl_sheet[f'H{start_row}'] = -(row_data['total_real10'])
                    pl_sheet[f'I{start_row}'] = -(row_data['total_real11'])
                    pl_sheet[f'J{start_row}'] = -(row_data['total_real12'])
                    pl_sheet[f'K{start_row}'] = -(row_data['total_real1'])
                    pl_sheet[f'L{start_row}'] = -(row_data['total_real2'])
                    pl_sheet[f'M{start_row}'] = -(row_data['total_real3'])
                    pl_sheet[f'N{start_row}'] = -(row_data['total_real4'])
                    pl_sheet[f'O{start_row}'] = -(row_data['total_real5'])
                    pl_sheet[f'P{start_row}'] = -(row_data['total_real6'])
                    pl_sheet[f'Q{start_row}'] = -(row_data['total_real7'])
                    pl_sheet[f'R{start_row}'] = -(row_data['total_real8'])
                    pl_sheet[f'T{start_row}'] = -(row_data['ytd_total'])             
                    pl_sheet[f'U{start_row}'].value = (row_data['variances'])

                    fpr_row_end = start_row


                    start_row += 1

        fpr_end = start_row

        if fpr_row_end is not None:
            for col in range(4, 22): 
                try:

                    cell = pl_sheet.cell(row=fpr_row_end, column=col)
                    cell.style = normal_cell_bottom_border
                except KeyError as e:
                    print(f"Error hiding row {col}: {e}") 
        for row in range(fpr_row_start, fpr_end):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True

            except KeyError as e:
                print(f"Error hiding row {row}: {e}") 
            # FEDERAL PROGRAM REVENUE TOTAL
        for col in range(2, 22):
            try:  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.font = fontbold

            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 

        for col in range(4, 22):  # Columns G to U
            try:

                cell = pl_sheet.cell(row=start_row, column=col)        
                cell.style = currency_style_noborder
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        pl_sheet[f'B{start_row}'] = 'Federal Program Revenue'
        pl_sheet[f'D{start_row}'] = totals["total_ammended_fpr"]    
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_fpr"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_fpr"]["09"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_fpr"]["10"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_fpr"]["11"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_fpr"]["12"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_fpr"]["01"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_fpr"]["02"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_fpr"]["03"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_fpr"]["04"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_fpr"]["05"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_fpr"]["06"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_fpr"]["07"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_fpr"]["08"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_fpr"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue_fpr"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'   
        start_row += 1

        total_revenue_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
    

            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Total Revenue'
        pl_sheet[f'D{start_row}'] = totals["total_ammended"]    
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_revenue"]["09"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_revenue"]["10"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_revenue"]["11"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_revenue"]["12"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_revenue"]["01"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_revenue"]["02"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_revenue"]["03"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_revenue"]["04"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_revenue"]["05"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_revenue"]["06"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_revenue"]["07"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_revenue"]["08"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_revenue"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'     

        start_row += 1   
        first_total_start = start_row
        first_total_end = None
        for row_data in data2: #1st TOTAL
            if row_data["category"] != 'Depreciation and Amortization':
                all_zeros = all(row_data[f'total_func{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'B{start_row}'] = f'{row_data["func_func"]} - {row_data["desc"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                    pl_sheet[f'G{start_row}'] = row_data['total_func9']
                    pl_sheet[f'H{start_row}'] = row_data['total_func10']
                    pl_sheet[f'I{start_row}'] = row_data['total_func11']
                    pl_sheet[f'J{start_row}'] = row_data['total_func12']
                    pl_sheet[f'K{start_row}'] = row_data['total_func1']
                    pl_sheet[f'L{start_row}'] = row_data['total_func2']
                    pl_sheet[f'M{start_row}'] = row_data['total_func3']
                    pl_sheet[f'N{start_row}'] = row_data['total_func4']
                    pl_sheet[f'O{start_row}'] = row_data['total_func5']
                    pl_sheet[f'P{start_row}'] = row_data['total_func6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func7']
                    pl_sheet[f'R{start_row}'] = row_data['total_func8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    pl_sheet[f'U{start_row}'] = row_data['variances']
                    pl_sheet[f'v{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")'
                    first_total_end = start_row
                    start_row += 1

        if first_total_end is not None:
            for row in range(first_total_start, first_total_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1


                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")   
    
        first_total_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)

            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Total'
        pl_sheet[f'D{start_row}'] = totals["first_total"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_total_first"] 
        pl_sheet[f'G{start_row}'] = totals["first_total_months"]["09"] 
        pl_sheet[f'H{start_row}'] = totals["first_total_months"]["10"] 
        pl_sheet[f'I{start_row}'] = totals["first_total_months"]["11"] 
        pl_sheet[f'J{start_row}'] = totals["first_total_months"]["12"] 
        pl_sheet[f'K{start_row}'] = totals["first_total_months"]["01"] 
        pl_sheet[f'L{start_row}'] = totals["first_total_months"]["02"] 
        pl_sheet[f'M{start_row}'] = totals["first_total_months"]["03"] 
        pl_sheet[f'N{start_row}'] = totals["first_total_months"]["04"] 
        pl_sheet[f'O{start_row}'] = totals["first_total_months"]["05"] 
        pl_sheet[f'P{start_row}'] = totals["first_total_months"]["06"] 
        pl_sheet[f'Q{start_row}'] = totals["first_total_months"]["07"] 
        pl_sheet[f'R{start_row}'] = totals["first_total_months"]["08"] 
        pl_sheet[f'T{start_row}'] = totals["first_ytd_total"]
        pl_sheet[f'U{start_row}'] = totals["variances_first_total"]

        pl_sheet[f'V{start_row}'].value = f'=IFERROR(+T{start_row}/E{start_row},"    ")'

        start_row += 2 #surplus (deficits) before depreciation
        surplus_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)

            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Surplus (Deficits) before Depreciation'
        pl_sheet[f'D{start_row}'] = totals["ammended_budget_SBD"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_SBD"] 
        pl_sheet[f'G{start_row}'] = totals["total_SBD"]["09"] 
        pl_sheet[f'H{start_row}'] = totals["total_SBD"]["10"] 
        pl_sheet[f'I{start_row}'] = totals["total_SBD"]["11"] 
        pl_sheet[f'J{start_row}'] = totals["total_SBD"]["12"] 
        pl_sheet[f'K{start_row}'] = totals["total_SBD"]["01"] 
        pl_sheet[f'L{start_row}'] = totals["total_SBD"]["02"] 
        pl_sheet[f'M{start_row}'] = totals["total_SBD"]["03"] 
        pl_sheet[f'N{start_row}'] = totals["total_SBD"]["04"] 
        pl_sheet[f'O{start_row}'] = totals["total_SBD"]["05"] 
        pl_sheet[f'P{start_row}'] = totals["total_SBD"]["06"] 
        pl_sheet[f'Q{start_row}'] = totals["total_SBD"]["07"] 
        pl_sheet[f'R{start_row}'] = totals["total_SBD"]["08"] 
        pl_sheet[f'T{start_row}'] = totals["ytd_SBD"] 
        pl_sheet[f'U{start_row}'] = totals["variances_SBD"] 
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 

        start_row += 2

        dna_row_start = start_row
        dna_row_end = None
        for row_data in data2: #Depreciation and amortization
            if row_data["category"] == 'Depreciation and Amortization':
                all_zeros = all(row_data[f'total_func2_{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'B{start_row}'] = f'{row_data["func_func"]} - {row_data["desc"]}'

                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']

                    pl_sheet[f'G{start_row}'] = row_data['total_func2_9']
                    pl_sheet[f'H{start_row}'] = row_data['total_func2_10']
                    pl_sheet[f'I{start_row}'] = row_data['total_func2_11']
                    pl_sheet[f'J{start_row}'] = row_data['total_func2_12']
                    pl_sheet[f'K{start_row}'] = row_data['total_func2_1']
                    pl_sheet[f'L{start_row}'] = row_data['total_func2_2']
                    pl_sheet[f'M{start_row}'] = row_data['total_func2_3']
                    pl_sheet[f'N{start_row}'] = row_data['total_func2_4']
                    pl_sheet[f'O{start_row}'] = row_data['total_func2_5']
                    pl_sheet[f'P{start_row}'] = row_data['total_func2_6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func2_7']
                    pl_sheet[f'R{start_row}'] = row_data['total_func2_8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    pl_sheet[f'U{start_row}'] = row_data['variances']
                    pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")'  
                    dna_row_end = start_row
                    start_row += 1


        dna_row = start_row
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)

            cell.style = currency_style
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        pl_sheet[f'B{start_row}'] = 'Depreciation and Amortization'
        pl_sheet[f'D{start_row}'] = totals["dna_total"]
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_dna"]
        pl_sheet[f'G{start_row}'] = totals["dna_total_months"]["09"] 
        pl_sheet[f'H{start_row}'] = totals["dna_total_months"]["10"] 
        pl_sheet[f'I{start_row}'] = totals["dna_total_months"]["11"] 
        pl_sheet[f'J{start_row}'] = totals["dna_total_months"]["12"] 
        pl_sheet[f'K{start_row}'] = totals["dna_total_months"]["01"] 
        pl_sheet[f'L{start_row}'] = totals["dna_total_months"]["02"] 
        pl_sheet[f'M{start_row}'] = totals["dna_total_months"]["03"] 
        pl_sheet[f'N{start_row}'] = totals["dna_total_months"]["04"] 
        pl_sheet[f'O{start_row}'] = totals["dna_total_months"]["05"] 
        pl_sheet[f'P{start_row}'] = totals["dna_total_months"]["06"] 
        pl_sheet[f'Q{start_row}'] = totals["dna_total_months"]["07"] 
        pl_sheet[f'R{start_row}'] = totals["dna_total_months"]["08"] 
        pl_sheet[f'T{start_row}'] = totals["dna_ytd_total"] 
        pl_sheet[f'U{start_row}'] = totals["variances_dna"]
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/E{start_row},"    ")' 

        start_row += 2
        netsurplus_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold

        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Net Surplus(Deficit)'
        pl_sheet[f'D{start_row}'] = totals["ammended_budget_netsurplus"]
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_netsurplus"]
        pl_sheet[f'G{start_row}'] = totals["total_netsurplus_months"]["09"]
        pl_sheet[f'H{start_row}'] = totals["total_netsurplus_months"]["10"]
        pl_sheet[f'I{start_row}'] = totals["total_netsurplus_months"]["11"]
        pl_sheet[f'J{start_row}'] = totals["total_netsurplus_months"]["12"]
        pl_sheet[f'K{start_row}'] = totals["total_netsurplus_months"]["01"]
        pl_sheet[f'L{start_row}'] = totals["total_netsurplus_months"]["02"]
        pl_sheet[f'M{start_row}'] = totals["total_netsurplus_months"]["03"]
        pl_sheet[f'N{start_row}'] = totals["total_netsurplus_months"]["04"]
        pl_sheet[f'O{start_row}'] = totals["total_netsurplus_months"]["05"]
        pl_sheet[f'P{start_row}'] = totals["total_netsurplus_months"]["06"]
        pl_sheet[f'Q{start_row}'] = totals["total_netsurplus_months"]["07"]
        pl_sheet[f'R{start_row}'] = totals["total_netsurplus_months"]["08"]
        pl_sheet[f'T{start_row}'] = totals["ytd_netsurplus"] 
        pl_sheet[f'U{start_row}'] = totals["variances_netsurplus"]
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")'   




        start_row += 2
        pl_sheet[f'B{start_row}'] = 'Expense By Object Codes'
        pl_sheet[f'B{start_row}'].font = fontbold

        start_row += 1
        payroll_row_start = start_row
        payroll_row_end = None 
        for row_data in data_activities: 
            if row_data['Category'] == 'Payroll and Benefits':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']

                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    payroll_row_end = start_row
                    start_row += 1

        if payroll_row_end is not None:
            for row in range(payroll_row_start, payroll_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")    

        payroll_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6100':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
  
                pl_sheet[f'D{start_row}'] = totals["total_budget_pc"]
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_pc"]
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pc"]["09"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pc"]["10"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pc"]["11"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pc"]["12"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pc"]["01"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pc"]["02"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pc"]["03"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pc"]["04"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pc"]["05"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pc"]["06"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pc"]["07"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pc"]["08"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_pc"]
                pl_sheet[f'U{start_row}'] = row_data['variances']
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1

        pcs_row_start = start_row
        pcs_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Professional and Contract Services':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    pcs_row_end = start_row
                    start_row += 1

        if pcs_row_end is not None:
            for row in range(pcs_row_start, pcs_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")    

        pcs_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6200':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_pcs"]
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_pcs"]
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pcs"]["09"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pcs"]["10"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pcs"]["11"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pcs"]["12"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pcs"]["01"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pcs"]["02"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pcs"]["03"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pcs"]["04"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pcs"]["05"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pcs"]["06"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pcs"]["07"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pcs"]["08"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_pcs"]

                pl_sheet[f'U{start_row}'] = row_data['variances'] 
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1

        sm_row_start = start_row
        sm_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Materials and Supplies':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    sm_row_end = start_row
                    start_row += 1

        if sm_row_end is not None:
            for row in range(sm_row_start, sm_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")   

        sm_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6300':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_sm"] 
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_sm"] 
                pl_sheet[f'G{start_row}'] = totals["total_EOC_sm"]["09"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_sm"]["10"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_sm"]["11"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_sm"]["12"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_sm"]["01"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_sm"]["02"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_sm"]["03"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_sm"]["04"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_sm"]["05"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_sm"]["06"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_sm"]["07"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_sm"]["08"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_sm"] 

                pl_sheet[f'U{start_row}'] = row_data['variances'] 
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1

        ooe_row_start = start_row
        ooe_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Other Operating Costs':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = totals["total_budget_sm"] 
                    pl_sheet[f'E{start_row}'] = totals["ytd_budget_sm"] 
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    ooe_row_end = start_row
                    start_row += 1

        if ooe_row_end is not None:
            for row in range(ooe_row_start, ooe_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")  

        ooe_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6400':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_ooe"]
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_ooe"]
                pl_sheet[f'G{start_row}'] = totals["total_EOC_ooe"]["09"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_ooe"]["10"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_ooe"]["11"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_ooe"]["12"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_ooe"]["01"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_ooe"]["02"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_ooe"]["03"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_ooe"]["04"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_ooe"]["05"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_ooe"]["06"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_ooe"]["07"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_ooe"]["08"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_ooe"]

                pl_sheet[f'U{start_row}'] = row_data['variances']  
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 

                start_row += 1




        oe_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6449':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["dna_total"]
                pl_sheet[f'E{start_row}'] = totals["ytd_ammended_dna"]
                pl_sheet[f'G{start_row}'] = totals["dna_total_months"]["09"]
                pl_sheet[f'H{start_row}'] = totals["dna_total_months"]["10"]
                pl_sheet[f'I{start_row}'] = totals["dna_total_months"]["11"]
                pl_sheet[f'J{start_row}'] = totals["dna_total_months"]["12"]
                pl_sheet[f'K{start_row}'] = totals["dna_total_months"]["01"]
                pl_sheet[f'L{start_row}'] = totals["dna_total_months"]["02"]
                pl_sheet[f'M{start_row}'] = totals["dna_total_months"]["03"]
                pl_sheet[f'N{start_row}'] = totals["dna_total_months"]["04"]
                pl_sheet[f'O{start_row}'] = totals["dna_total_months"]["05"]
                pl_sheet[f'P{start_row}'] = totals["dna_total_months"]["06"]
                pl_sheet[f'Q{start_row}'] = totals["dna_total_months"]["07"]
                pl_sheet[f'R{start_row}'] = totals["dna_total_months"]["08"]
                pl_sheet[f'T{start_row}'] = totals["dna_ytd_total"]

                pl_sheet[f'U{start_row}'] = totals["variances_dna"]
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 

                start_row += 1

        total_expense_row_end = None
        total_expense_row_start = start_row
        
        for row_data in data_activities: 
            if row_data['Category'] == 'Debt Services':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data["total_budget"]
                    pl_sheet[f'E{start_row}'] = row_data["ytd_budget"]
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
    
    
                    total_expense_row_end = start_row
                    start_row += 1
        if total_expense_row_end is not None:
            for row in range(total_expense_row_start, total_expense_row_end+1):
                try:
                    
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")  

        total_expense_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6500':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_te"]
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_te"]
                pl_sheet[f'G{start_row}'] = totals["total_EOC_te"]["09"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_te"]["10"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_te"]["11"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_te"]["12"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_te"]["01"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_te"]["02"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_te"]["03"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_te"]["04"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_te"]["05"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_te"]["06"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_te"]["07"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_te"]["08"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_te"]

                pl_sheet[f'U{start_row}'] = row_data['variances'] 
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1

        start_row += 1

        total_expense_total = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Total Expense'
        pl_sheet[f'D{start_row}'] = totals["total_expense"]
        pl_sheet[f'E{start_row}'] = totals["total_expense_ytd_budget"] 
        pl_sheet[f'G{start_row}'] = totals["total_expense_months"]["09"]
        pl_sheet[f'H{start_row}'] = totals["total_expense_months"]["10"]
        pl_sheet[f'I{start_row}'] = totals["total_expense_months"]["11"]
        pl_sheet[f'J{start_row}'] = totals["total_expense_months"]["12"]
        pl_sheet[f'K{start_row}'] = totals["total_expense_months"]["01"]
        pl_sheet[f'L{start_row}'] = totals["total_expense_months"]["02"]
        pl_sheet[f'M{start_row}'] = totals["total_expense_months"]["03"]
        pl_sheet[f'N{start_row}'] = totals["total_expense_months"]["04"]
        pl_sheet[f'O{start_row}'] = totals["total_expense_months"]["05"]
        pl_sheet[f'P{start_row}'] = totals["total_expense_months"]["06"]
        pl_sheet[f'Q{start_row}'] = totals["total_expense_months"]["07"]
        pl_sheet[f'R{start_row}'] = totals["total_expense_months"]["08"]

        pl_sheet[f'T{start_row}'] = totals["total_expense_ytd"] 
        pl_sheet[f'U{start_row}'] = totals["variances_total_expense"]
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 


        start_row += 1
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Net Income'
        pl_sheet[f'D{start_row}'] = totals["budget_net_income"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_budget_net_income"] 
        pl_sheet[f'G{start_row}'] = totals["total_net_income_months"]["09"]
        pl_sheet[f'H{start_row}'] = totals["total_net_income_months"]["10"]
        pl_sheet[f'I{start_row}'] = totals["total_net_income_months"]["11"]
        pl_sheet[f'J{start_row}'] = totals["total_net_income_months"]["12"]
        pl_sheet[f'K{start_row}'] = totals["total_net_income_months"]["01"]
        pl_sheet[f'L{start_row}'] = totals["total_net_income_months"]["02"]
        pl_sheet[f'M{start_row}'] = totals["total_net_income_months"]["03"]
        pl_sheet[f'N{start_row}'] = totals["total_net_income_months"]["04"]
        pl_sheet[f'O{start_row}'] = totals["total_net_income_months"]["05"]
        pl_sheet[f'P{start_row}'] = totals["total_net_income_months"]["06"]
        pl_sheet[f'Q{start_row}'] = totals["total_net_income_months"]["07"]
        pl_sheet[f'R{start_row}'] = totals["total_net_income_months"]["08"]

        pl_sheet[f'T{start_row}'] = totals["ytd_net_income"] 
        pl_sheet[f'U{start_row}'] = totals["variances_net_income"] 
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 



        start_row += 4 #Total expense and Net income





    # FOR PROFIT_LOSS OF MANARA AND PREPSCHOOL   
    else:
        for col in range(7, 20 ):
            col_letter = get_column_letter(col)
            pl_sheet.column_dimensions[col_letter].outline_level = 1
            pl_sheet.column_dimensions[col_letter].hidden = True

        last_number = months["last_month_number"]
        # PL START OF DESIGN
        if last_number <= 6:
            last_number += 13
        else:
            last_number += 1

     
        for col in range(last_number,19):
            col_letter = get_column_letter(col)     
            pl_sheet.column_dimensions[col_letter].outline_level = 2
            pl_sheet.column_dimensions[col_letter].hidden = True
           
        start_pl = 1
        pl_sheet[f'B{start_pl}'] = f'{school_name}\nFY2022-2023 Statement of\nActivities as of {months["last_month"]}'
        start_pl += 2
        pl_sheet[f'E{start_pl}'] = f'{months["format_ytd_budget"]}% YTD \nBUDGET'
        pl_sheet[f'G{start_pl}'] = 'July'
        pl_sheet[f'H{start_pl}'] = 'August'
        pl_sheet[f'I{start_pl}'] = 'September'
        pl_sheet[f'J{start_pl}'] = 'October'
        pl_sheet[f'K{start_pl}'] = 'November'
        pl_sheet[f'L{start_pl}'] = 'December'
        pl_sheet[f'M{start_pl}'] = 'January'
        pl_sheet[f'N{start_pl}'] = 'February'
        pl_sheet[f'O{start_pl}'] = 'March'
        pl_sheet[f'P{start_pl}'] = 'April'
        pl_sheet[f'Q{start_pl}'] = 'May'
        pl_sheet[f'R{start_pl}'] = 'June'
        pl_sheet[f'V{start_pl}'] = f'Var. {months["format_ytd_budget"]}'
        start_row = 5
        lr_row_start = start_row
        lr_row_end = None
        for row_data in data:
            if row_data['category'] == 'Local Revenue':
                all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'A{start_row}'] = row_data['fund']
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 

                    pl_sheet[f'G{start_row}'] = -(row_data['total_real7'])
                    pl_sheet[f'H{start_row}'] = -(row_data['total_real8'])
                    pl_sheet[f'I{start_row}'] = -(row_data['total_real9'])
                    pl_sheet[f'J{start_row}'] = -(row_data['total_real10'])
                    pl_sheet[f'K{start_row}'] = -(row_data['total_real11'])
                    pl_sheet[f'L{start_row}'] = -(row_data['total_real12'])
                    pl_sheet[f'M{start_row}'] = -(row_data['total_real1'])
                    pl_sheet[f'N{start_row}'] = -(row_data['total_real2'])
                    pl_sheet[f'O{start_row}'] = -(row_data['total_real3'])
                    pl_sheet[f'P{start_row}'] = -(row_data['total_real4'])
                    pl_sheet[f'Q{start_row}'] = -(row_data['total_real5'])
                    pl_sheet[f'R{start_row}'] = -(row_data['total_real6'])
                    pl_sheet[f'T{start_row}'] = -(row_data['ytd_total'])
                    lr_row_end = start_row



                    start_row += 1
       
        if lr_row_end is not None:
            for col in range(4, 22):
                try:  # Columns G to U
                    cell = pl_sheet.cell(row=lr_row_end, column=col)

                    cell.style = normal_cell_bottom_border
                except KeyError as e:
                    print(f"Error hiding row {col}: {e}")           
        
            for row in range(lr_row_start, lr_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True

                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")           
        lr_end = start_row
        #local revenue total
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)
            
            cell.style = currency_style_noborder
        pl_sheet[f'B{start_row}'] = 'Local Revenue'
        pl_sheet[f'D{start_row}'] =  totals["total_ammended_lr"]
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_lr"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_lr"]["07"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_lr"]["08"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_lr"]["09"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_lr"]["10"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_lr"]["11"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_lr"]["12"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_lr"]["01"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_lr"]["02"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_lr"]["03"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_lr"]["04"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_lr"]["05"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_lr"]["06"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_lr"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue_lr"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'
        
        
        start_row += 1  
    
        
    
        spr_row_start = start_row
        spr_row_end = None
        for row_data in data:
            if row_data['category'] == 'State Program Revenue':
                all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell

                    pl_sheet[f'A{start_row}'] = row_data['fund']
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 

                    pl_sheet[f'G{start_row}'] = -(row_data['total_real7'])
                    pl_sheet[f'H{start_row}'] = -(row_data['total_real8'])
                    pl_sheet[f'I{start_row}'] = -(row_data['total_real9'])
                    pl_sheet[f'J{start_row}'] = -(row_data['total_real10'])
                    pl_sheet[f'K{start_row}'] = -(row_data['total_real11'])
                    pl_sheet[f'L{start_row}'] = -(row_data['total_real12'])
                    pl_sheet[f'M{start_row}'] = -(row_data['total_real1'])
                    pl_sheet[f'N{start_row}'] = -(row_data['total_real2'])
                    pl_sheet[f'O{start_row}'] = -(row_data['total_real3'])
                    pl_sheet[f'P{start_row}'] = -(row_data['total_real4'])
                    pl_sheet[f'Q{start_row}'] = -(row_data['total_real5'])
                    pl_sheet[f'R{start_row}'] = -(row_data['total_real6'])
                    pl_sheet[f'T{start_row}'] = -(row_data['ytd_total'])

                    pl_sheet[f'U{start_row}']= (row_data['variances'])
                    spr_row_end = start_row



                    start_row += 1
    
        if spr_row_end is not None:
            for col in range(4, 22):  # Columns G to U
                cell = pl_sheet.cell(row=spr_row_end, column=col)

                cell.style = normal_cell_bottom_border

            for row in range(spr_row_start, spr_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True

                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")    
                
               
        spr_end = start_row
        # STATE PROGRAM TOTAL      
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)        
            cell.style = currency_style_noborder
        pl_sheet[f'B{start_row}'] = 'State Program Revenue'
        pl_sheet[f'D{start_row}'] =  totals["total_ammended_spr"]    
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_spr"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_spr"]["07"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_spr"]["08"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_spr"]["09"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_spr"]["10"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_spr"]["11"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_spr"]["12"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_spr"]["01"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_spr"]["02"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_spr"]["03"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_spr"]["04"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_spr"]["05"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_spr"]["06"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_spr"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue_spr"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'
    
    
        start_row += 1
        
        
        fpr_row_start = start_row
        fpr_row_end = None
        for row_data in data:
            if row_data['category'] == 'Federal Program Revenue':
                all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 

                    pl_sheet[f'A{start_row}'] = row_data['fund']
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']             
                    pl_sheet[f'G{start_row}'] = -(row_data['total_real7'])
                    pl_sheet[f'H{start_row}'] = -(row_data['total_real8'])
                    pl_sheet[f'I{start_row}'] = -(row_data['total_real9'])
                    pl_sheet[f'J{start_row}'] = -(row_data['total_real10'])
                    pl_sheet[f'K{start_row}'] = -(row_data['total_real11'])
                    pl_sheet[f'L{start_row}'] = -(row_data['total_real12'])
                    pl_sheet[f'M{start_row}'] = -(row_data['total_real1'])
                    pl_sheet[f'N{start_row}'] = -(row_data['total_real2'])
                    pl_sheet[f'O{start_row}'] = -(row_data['total_real3'])
                    pl_sheet[f'P{start_row}'] = -(row_data['total_real4'])
                    pl_sheet[f'Q{start_row}'] = -(row_data['total_real5'])
                    pl_sheet[f'R{start_row}'] = -(row_data['total_real6'])
                    pl_sheet[f'T{start_row}'] = -(row_data['ytd_total'])             
                    pl_sheet[f'U{start_row}'].value = (row_data['variances'])

                    fpr_row_end = start_row


                    start_row += 1
                
        fpr_end = start_row
        
        if fpr_row_end is not None:
            for col in range(4, 22):  # Columns G to U
                cell = pl_sheet.cell(row=fpr_row_end, column=col)

                cell.style = normal_cell_bottom_border

            for row in range(fpr_row_start, fpr_end):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}") 
            # FEDERAL PROGRAM REVENUE TOTAL
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)        
            cell.style = currency_style_noborder
        pl_sheet[f'B{start_row}'] = 'Federal Program Revenue'
        pl_sheet[f'D{start_row}'] = totals["total_ammended_fpr"]    
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_fpr"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_fpr"]["07"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_fpr"]["08"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_fpr"]["09"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_fpr"]["10"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_fpr"]["11"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_fpr"]["12"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_fpr"]["01"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_fpr"]["02"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_fpr"]["03"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_fpr"]["04"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_fpr"]["05"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_fpr"]["06"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_fpr"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue_fpr"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'   
        start_row += 1
    
        total_revenue_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
    
        
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Total Revenue'
        pl_sheet[f'D{start_row}'] = totals["total_ammended"]    
        pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total"]
        pl_sheet[f'G{start_row}'] =  -(totals["total_revenue"]["07"])
        pl_sheet[f'H{start_row}'] =  -(totals["total_revenue"]["08"])
        pl_sheet[f'I{start_row}'] =  -(totals["total_revenue"]["09"])
        pl_sheet[f'J{start_row}'] =  -(totals["total_revenue"]["10"])
        pl_sheet[f'K{start_row}'] =  -(totals["total_revenue"]["11"])
        pl_sheet[f'L{start_row}'] =  -(totals["total_revenue"]["12"])
        pl_sheet[f'M{start_row}'] =  -(totals["total_revenue"]["01"])
        pl_sheet[f'N{start_row}'] =  -(totals["total_revenue"]["02"])
        pl_sheet[f'O{start_row}'] =  -(totals["total_revenue"]["03"])
        pl_sheet[f'P{start_row}'] =  -(totals["total_revenue"]["04"])
        pl_sheet[f'Q{start_row}'] =  -(totals["total_revenue"]["05"])
        pl_sheet[f'R{start_row}'] =  -(totals["total_revenue"]["06"])
        pl_sheet[f'T{start_row}'] =  -(totals["ytd_total_revenue"])
        pl_sheet[f'U{start_row}'] =  totals["variances_revenue"]
        pl_sheet[f'V{start_row}'].value = f'=+T{start_row}/D{start_row}'     
    
        start_row += 1   
        first_total_start = start_row
        first_total_end = None
    
        for row_data in data2: #1st TOTAL
            if row_data["category"] != 'Depreciation and Amortization':
                all_zeros = all(row_data[f'total_func{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'B{start_row}'] = f'{row_data["func_func"]} - {row_data["desc"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                    pl_sheet[f'G{start_row}'] = row_data['total_func7']
                    pl_sheet[f'H{start_row}'] = row_data['total_func8']
                    pl_sheet[f'I{start_row}'] = row_data['total_func9']
                    pl_sheet[f'J{start_row}'] = row_data['total_func10']
                    pl_sheet[f'K{start_row}'] = row_data['total_func11']
                    pl_sheet[f'L{start_row}'] = row_data['total_func12']
                    pl_sheet[f'M{start_row}'] = row_data['total_func1']
                    pl_sheet[f'N{start_row}'] = row_data['total_func2']
                    pl_sheet[f'O{start_row}'] = row_data['total_func3']
                    pl_sheet[f'P{start_row}'] = row_data['total_func4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func5']
                    pl_sheet[f'R{start_row}'] = row_data['total_func6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    pl_sheet[f'U{start_row}'] = row_data['variances']
                    pl_sheet[f'v{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")'
                    first_total_end = start_row
                    start_row += 1


        if first_total_end is not None:
            for row in range(first_total_start, first_total_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1


                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")   
    
        first_total_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)
            
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Total'
        pl_sheet[f'D{start_row}'] = totals["first_total"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_total_first"] 
        pl_sheet[f'G{start_row}'] = totals["first_total_months"]["07"] 
        pl_sheet[f'H{start_row}'] = totals["first_total_months"]["08"] 
        pl_sheet[f'I{start_row}'] = totals["first_total_months"]["09"] 
        pl_sheet[f'J{start_row}'] = totals["first_total_months"]["10"] 
        pl_sheet[f'K{start_row}'] = totals["first_total_months"]["11"] 
        pl_sheet[f'L{start_row}'] = totals["first_total_months"]["12"] 
        pl_sheet[f'M{start_row}'] = totals["first_total_months"]["01"] 
        pl_sheet[f'N{start_row}'] = totals["first_total_months"]["02"] 
        pl_sheet[f'O{start_row}'] = totals["first_total_months"]["03"] 
        pl_sheet[f'P{start_row}'] = totals["first_total_months"]["04"] 
        pl_sheet[f'Q{start_row}'] = totals["first_total_months"]["05"] 
        pl_sheet[f'R{start_row}'] = totals["first_total_months"]["06"] 
        pl_sheet[f'T{start_row}'] = totals["first_ytd_total"]
        pl_sheet[f'U{start_row}'] = totals["variances_first_total"]
    
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(+T{start_row}/E{start_row},"    ")'
        
        start_row += 2 #surplus (deficits) before depreciation
        surplus_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)
            
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Surplus (Deficits) before Depreciation'
        pl_sheet[f'D{start_row}'] = totals["ammended_budget_SBD"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_SBD"] 
        pl_sheet[f'G{start_row}'] = totals["total_SBD"]["07"] 
        pl_sheet[f'H{start_row}'] = totals["total_SBD"]["08"] 
        pl_sheet[f'I{start_row}'] = totals["total_SBD"]["09"] 
        pl_sheet[f'J{start_row}'] = totals["total_SBD"]["10"] 
        pl_sheet[f'K{start_row}'] = totals["total_SBD"]["11"] 
        pl_sheet[f'L{start_row}'] = totals["total_SBD"]["12"] 
        pl_sheet[f'M{start_row}'] = totals["total_SBD"]["01"] 
        pl_sheet[f'N{start_row}'] = totals["total_SBD"]["02"] 
        pl_sheet[f'O{start_row}'] = totals["total_SBD"]["03"] 
        pl_sheet[f'P{start_row}'] = totals["total_SBD"]["04"] 
        pl_sheet[f'Q{start_row}'] = totals["total_SBD"]["05"] 
        pl_sheet[f'R{start_row}'] = totals["total_SBD"]["06"] 
        pl_sheet[f'T{start_row}'] = totals["ytd_SBD"] 
        pl_sheet[f'U{start_row}'] = totals["variances_SBD"] 
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
    
        start_row += 2
    
        dna_row_start = start_row
        dna_row_end = None
        for row_data in data2: #Depreciation and amortization
            if row_data["category"] == 'Depreciation and Amortization':
                all_zeros = all(row_data[f'total_func2_{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'B{start_row}'] = f'{row_data["func_func"]} - {row_data["desc"]}'

                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']

                    pl_sheet[f'G{start_row}'] = row_data['total_func2_7']
                    pl_sheet[f'H{start_row}'] = row_data['total_func2_8']
                    pl_sheet[f'I{start_row}'] = row_data['total_func2_9']
                    pl_sheet[f'J{start_row}'] = row_data['total_func2_10']
                    pl_sheet[f'K{start_row}'] = row_data['total_func2_11']
                    pl_sheet[f'L{start_row}'] = row_data['total_func2_12']
                    pl_sheet[f'M{start_row}'] = row_data['total_func2_1']
                    pl_sheet[f'N{start_row}'] = row_data['total_func2_2']
                    pl_sheet[f'O{start_row}'] = row_data['total_func2_3']
                    pl_sheet[f'P{start_row}'] = row_data['total_func2_4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func2_5']
                    pl_sheet[f'R{start_row}'] = row_data['total_func2_6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    pl_sheet[f'U{start_row}'] = row_data['variances']
                    pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")'  
                    dna_row_end = start_row
                    start_row += 1
    
    
        dna_row = start_row
        
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)
            
            cell.style = currency_style
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        pl_sheet[f'B{start_row}'] = 'Depreciation and Amortization'
        pl_sheet[f'D{start_row}'] = totals["dna_total"]
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_dna"]
        pl_sheet[f'G{start_row}'] = totals["dna_total_months"]["07"] 
        pl_sheet[f'H{start_row}'] = totals["dna_total_months"]["08"] 
        pl_sheet[f'I{start_row}'] = totals["dna_total_months"]["09"] 
        pl_sheet[f'J{start_row}'] = totals["dna_total_months"]["10"] 
        pl_sheet[f'K{start_row}'] = totals["dna_total_months"]["11"] 
        pl_sheet[f'L{start_row}'] = totals["dna_total_months"]["12"] 
        pl_sheet[f'M{start_row}'] = totals["dna_total_months"]["01"] 
        pl_sheet[f'N{start_row}'] = totals["dna_total_months"]["02"] 
        pl_sheet[f'O{start_row}'] = totals["dna_total_months"]["03"] 
        pl_sheet[f'P{start_row}'] = totals["dna_total_months"]["04"] 
        pl_sheet[f'Q{start_row}'] = totals["dna_total_months"]["05"] 
        pl_sheet[f'R{start_row}'] = totals["dna_total_months"]["06"] 
        pl_sheet[f'T{start_row}'] = totals["dna_ytd_total"] 
        pl_sheet[f'U{start_row}'] = totals["variances_dna"]
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/E{start_row},"    ")' 
    
        start_row += 2
        netsurplus_row = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
    
        for col in range(4, 22):  # Columns D to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Net Surplus(Deficit)'
        pl_sheet[f'D{start_row}'] = totals["ammended_budget_netsurplus"]
        pl_sheet[f'E{start_row}'] = totals["ytd_ammended_netsurplus"]
        pl_sheet[f'G{start_row}'] = totals["total_netsurplus_months"]["07"]
        pl_sheet[f'H{start_row}'] = totals["total_netsurplus_months"]["08"]
        pl_sheet[f'I{start_row}'] = totals["total_netsurplus_months"]["09"]
        pl_sheet[f'J{start_row}'] = totals["total_netsurplus_months"]["10"]
        pl_sheet[f'K{start_row}'] = totals["total_netsurplus_months"]["11"]
        pl_sheet[f'L{start_row}'] = totals["total_netsurplus_months"]["12"]
        pl_sheet[f'M{start_row}'] = totals["total_netsurplus_months"]["01"]
        pl_sheet[f'N{start_row}'] = totals["total_netsurplus_months"]["02"]
        pl_sheet[f'O{start_row}'] = totals["total_netsurplus_months"]["03"]
        pl_sheet[f'P{start_row}'] = totals["total_netsurplus_months"]["04"]
        pl_sheet[f'Q{start_row}'] = totals["total_netsurplus_months"]["05"]
        pl_sheet[f'R{start_row}'] = totals["total_netsurplus_months"]["06"]
        pl_sheet[f'T{start_row}'] = totals["ytd_netsurplus"] 
        pl_sheet[f'U{start_row}'] = totals["variances_netsurplus"]
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")'   
        
    
    
    
        start_row += 2
        pl_sheet[f'B{start_row}'] = 'Expense By Object Codes'
        pl_sheet[f'B{start_row}'].font = fontbold
    
        start_row += 1
        payroll_row_start = start_row
        payroll_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Payroll and Benefits':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  # Columns G to U
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell 
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'

                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']

                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    payroll_row_end = start_row
                    start_row += 1


        if payroll_row_end is not None:

            for row in range(payroll_row_start, payroll_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")    
    
        payroll_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6100':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_pc"]
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_pc"]
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pc"]["07"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pc"]["08"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pc"]["09"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pc"]["10"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pc"]["11"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pc"]["12"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pc"]["01"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pc"]["02"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pc"]["03"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pc"]["04"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pc"]["05"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pc"]["06"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_pc"]
                pl_sheet[f'U{start_row}'] = row_data['variances']
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1
    
        pcs_row_start = start_row
        pcs_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Professional and Contract Services':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    pcs_row_end = start_row
                    start_row += 1
    

        if pcs_row_end is not None:
            for row in range(pcs_row_start, pcs_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")    
    
        pcs_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6200':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_pcs"]
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_pcs"]
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pcs"]["07"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pcs"]["08"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pcs"]["09"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pcs"]["10"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pcs"]["11"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pcs"]["12"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pcs"]["01"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pcs"]["02"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pcs"]["03"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pcs"]["04"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pcs"]["05"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pcs"]["06"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_pcs"]
               
                pl_sheet[f'U{start_row}'] = row_data['variances'] 
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1
    
        sm_row_start = start_row
        sm_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Materials and Supplies':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    sm_row_end = start_row
                    start_row += 1
    

        if sm_row_end is not None:
            for row in range(sm_row_start, sm_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")   

        sm_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6300':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals['total_budget_sm']
                pl_sheet[f'E{start_row}'] = totals['ytd_budget_sm'] 
                pl_sheet[f'G{start_row}'] = totals["total_EOC_sm"]["07"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_sm"]["08"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_sm"]["09"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_sm"]["10"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_sm"]["11"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_sm"]["12"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_sm"]["01"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_sm"]["02"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_sm"]["03"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_sm"]["04"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_sm"]["05"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_sm"]["06"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_sm"] 
                
                pl_sheet[f'U{start_row}'] = row_data['variances'] 
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1
            
        ooe_row_start = start_row
        ooe_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Other Operating Costs':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                    ooe_row_end = start_row
                    start_row += 1
    
        if ooe_row_end is not None:
            for row in range(ooe_row_start, ooe_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")  
    
        ooe_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6400':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals['total_budget_ooe']
                pl_sheet[f'E{start_row}'] = totals['ytd_budget_ooe'] 
                pl_sheet[f'G{start_row}'] = totals["total_EOC_ooe"]["07"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_ooe"]["08"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_ooe"]["09"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_ooe"]["10"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_ooe"]["11"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_ooe"]["12"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_ooe"]["01"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_ooe"]["02"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_ooe"]["03"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_ooe"]["04"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_ooe"]["05"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_ooe"]["06"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_ooe"]
               
                pl_sheet[f'U{start_row}'] = row_data['variances']  
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                
                start_row += 1
    
        
    
        oe_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6449':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals['dna_total']
                pl_sheet[f'E{start_row}'] = totals['ytd_ammended_dna'] 
                pl_sheet[f'G{start_row}'] = totals["dna_total_months"]["07"]
                pl_sheet[f'H{start_row}'] = totals["dna_total_months"]["08"]
                pl_sheet[f'I{start_row}'] = totals["dna_total_months"]["09"]
                pl_sheet[f'J{start_row}'] = totals["dna_total_months"]["10"]
                pl_sheet[f'K{start_row}'] = totals["dna_total_months"]["11"]
                pl_sheet[f'L{start_row}'] = totals["dna_total_months"]["12"]
                pl_sheet[f'M{start_row}'] = totals["dna_total_months"]["01"]
                pl_sheet[f'N{start_row}'] = totals["dna_total_months"]["02"]
                pl_sheet[f'O{start_row}'] = totals["dna_total_months"]["03"]
                pl_sheet[f'P{start_row}'] = totals["dna_total_months"]["04"]
                pl_sheet[f'Q{start_row}'] = totals["dna_total_months"]["05"]
                pl_sheet[f'R{start_row}'] = totals["dna_total_months"]["06"]
                pl_sheet[f'T{start_row}'] = totals["dna_ytd_total"]               
                pl_sheet[f'U{start_row}'] = totals['variances_dna']  
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                
                start_row += 1
    
    
        total_expense_row_start = start_row
        total_expense_row_end = None
        for row_data in data_activities: 
            if row_data['Category'] == 'Debt Services':
                all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    for col in range(4, 22):  
                        cell = pl_sheet.cell(row=start_row, column=col)
                        cell.style = normal_cell
                    pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                    pl_sheet[f'D{start_row}'] = row_data['total_budget']
                    pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'T{start_row}'] = row_data['ytd_total']


                    total_expense_row_end = start_row
                    start_row += 1
    
        if total_expense_row_end is not None:
            for row in range(total_expense_row_start, total_expense_row_end+1):
                try:
                    pl_sheet.row_dimensions[row].outline_level = 1
                    pl_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")  
    
        total_expense_row = start_row
        for row_data in data_expensebyobject: 
            if row_data['obj'] == '6500':
                for col in range(4, 22):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals['total_budget_te']
                pl_sheet[f'E{start_row}'] = totals['ytd_budget_te'] 
                pl_sheet[f'G{start_row}'] = totals["total_EOC_te"]["07"]
                pl_sheet[f'H{start_row}'] = totals["total_EOC_te"]["08"]
                pl_sheet[f'I{start_row}'] = totals["total_EOC_te"]["09"]
                pl_sheet[f'J{start_row}'] = totals["total_EOC_te"]["10"]
                pl_sheet[f'K{start_row}'] = totals["total_EOC_te"]["11"]
                pl_sheet[f'L{start_row}'] = totals["total_EOC_te"]["12"]
                pl_sheet[f'M{start_row}'] = totals["total_EOC_te"]["01"]
                pl_sheet[f'N{start_row}'] = totals["total_EOC_te"]["02"]
                pl_sheet[f'O{start_row}'] = totals["total_EOC_te"]["03"]
                pl_sheet[f'P{start_row}'] = totals["total_EOC_te"]["04"]
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_te"]["05"]
                pl_sheet[f'R{start_row}'] = totals["total_EOC_te"]["06"]
                pl_sheet[f'T{start_row}'] = totals["ytd_EOC_te"]
                
                pl_sheet[f'U{start_row}'] = row_data['variances'] 
                pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
                start_row += 1
    
        start_row += 1
    
        total_expense_total = start_row
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Total Expense'
        pl_sheet[f'D{start_row}'] = totals["total_expense"]
        pl_sheet[f'E{start_row}'] = totals["total_expense_ytd_budget"] 
        pl_sheet[f'G{start_row}'] = totals["total_expense_months"]["07"]
        pl_sheet[f'H{start_row}'] = totals["total_expense_months"]["08"]
        pl_sheet[f'I{start_row}'] = totals["total_expense_months"]["09"]
        pl_sheet[f'J{start_row}'] = totals["total_expense_months"]["10"]
        pl_sheet[f'K{start_row}'] = totals["total_expense_months"]["11"]
        pl_sheet[f'L{start_row}'] = totals["total_expense_months"]["12"]
        pl_sheet[f'M{start_row}'] = totals["total_expense_months"]["01"]
        pl_sheet[f'N{start_row}'] = totals["total_expense_months"]["02"]
        pl_sheet[f'O{start_row}'] = totals["total_expense_months"]["03"]
        pl_sheet[f'P{start_row}'] = totals["total_expense_months"]["04"]
        pl_sheet[f'Q{start_row}'] = totals["total_expense_months"]["05"]
        pl_sheet[f'R{start_row}'] = totals["total_expense_months"]["06"]
         
        pl_sheet[f'T{start_row}'] = totals["total_expense_ytd"] 
        pl_sheet[f'U{start_row}'] = totals["variances_total_expense"]
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
        
    
        start_row += 1
        for col in range(2, 22):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 22):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Net Income'
        pl_sheet[f'D{start_row}'] = totals["budget_net_income"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_budget_net_income"] 
        pl_sheet[f'G{start_row}'] = totals["total_net_income_months"]["07"]
        pl_sheet[f'H{start_row}'] = totals["total_net_income_months"]["08"]
        pl_sheet[f'I{start_row}'] = totals["total_net_income_months"]["09"]
        pl_sheet[f'J{start_row}'] = totals["total_net_income_months"]["10"]
        pl_sheet[f'K{start_row}'] = totals["total_net_income_months"]["11"]
        pl_sheet[f'L{start_row}'] = totals["total_net_income_months"]["12"]
        pl_sheet[f'M{start_row}'] = totals["total_net_income_months"]["01"]
        pl_sheet[f'N{start_row}'] = totals["total_net_income_months"]["02"]
        pl_sheet[f'O{start_row}'] = totals["total_net_income_months"]["03"]
        pl_sheet[f'P{start_row}'] = totals["total_net_income_months"]["04"]
        pl_sheet[f'Q{start_row}'] = totals["total_net_income_months"]["05"]
        pl_sheet[f'R{start_row}'] = totals["total_net_income_months"]["06"]
        
        pl_sheet[f'T{start_row}'] = totals["ytd_net_income"] 
        pl_sheet[f'U{start_row}'] = totals["variances_net_income"] 
        pl_sheet[f'V{start_row}'].value = f'=IFERROR(T{start_row}/D{start_row},"    ")' 
        
    
    
        start_row += 4 #Total expense and Net income
 
    #BS DESIGN

    for row in range(2,200):
        bs_sheet.row_dimensions[row].height = 19
    # bs_sheet.row_dimensions[17].height = 26 #local revenue
    # bs_sheet.row_dimensions[20].height = 26 #spr
    # bs_sheet.row_dimensions[33].height = 26 #fpr

    
    
    bs_sheet.column_dimensions['A'].width = 8
    bs_sheet.column_dimensions['B'].width = 32
    bs_sheet.column_dimensions['C'].hidden = True
    bs_sheet.column_dimensions['D'].width = 14
    bs_sheet.column_dimensions['E'].width = 28
    bs_sheet.column_dimensions['F'].width = 15
    bs_sheet.column_dimensions['G'].width = 15
    bs_sheet.column_dimensions['H'].width = 13
    bs_sheet.column_dimensions['I'].width = 13
    bs_sheet.column_dimensions['J'].width = 13
    bs_sheet.column_dimensions['K'].width = 13
    bs_sheet.column_dimensions['L'].width = 13
    bs_sheet.column_dimensions['M'].width = 13
    bs_sheet.column_dimensions['N'].width = 13
    bs_sheet.column_dimensions['O'].width = 13
    bs_sheet.column_dimensions['P'].width = 13
    bs_sheet.column_dimensions['Q'].width = 13
    bs_sheet.column_dimensions['R'].width = 13
    bs_sheet.column_dimensions['S'].width = 3
    bs_sheet.column_dimensions['T'].width = 17
    bs_sheet.column_dimensions['U'].width = 17
    bs_sheet.column_dimensions['V'].width = 12

    indent_style = NamedStyle(name="indent_style", alignment=Alignment(indent=2))
    indent_style2 = NamedStyle(name="indent_style2", alignment=Alignment(indent=4))

    start_bs = 1
    bs_sheet[f'D{start_bs}'] = f'{school_name}\nFY{months["FY_year_1"]}-{months["FY_year_2"]} Balance Sheet as of {months["last_month"]}'
    #--- BS INSERT
    header_bs = 3
    bs_sheet[f'U{header_bs}'] = f'As of {months["last_month_name"]}'


    if  school != 'manara' and school != 'leadership':
        
        
        for col in range(7, 20 ):
            col_letter = get_column_letter(col)
            bs_sheet.column_dimensions[col_letter].outline_level = 1
            bs_sheet.column_dimensions[col_letter].hidden = True

        last_number = months["last_month_number"]
        # PL START OF DESIGN
        if last_number <= 8:
            last_number += 11
        else:
            last_number -= 1


        for col in range(last_number,19):
            col_letter = get_column_letter(col)

      
            bs_sheet.column_dimensions[col_letter].outline_level = 2
            bs_sheet.column_dimensions[col_letter].hidden = True

        acc_per = months["last_month_number"]
        
        if acc_per >= 10:
            acc_per = str(acc_per)
        else:
            acc_per = str(f'0{acc_per}')



        start_bs_for_hiding = 7
        start_row_bs = 6
        hide_row_bs_start = start_row_bs
        hide_row_bs_end= None
        bs_sheet[f'D{start_row_bs}'] = 'Current Assets'
        for row in data_activitybs:
            if row['Activity'] == 'Cash':
      
                start_row_bs += 1
                hide_row_bs_end = start_row_bs
                for col in range(5, 22):  # Columns G to U
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell 
                bs_sheet[f'D{start_row_bs}'].style = indent_style
                bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                bs_sheet[f'G{start_row_bs}'] =  row['total_bal9']
                bs_sheet[f'H{start_row_bs}'] =  row['total_bal10']
                bs_sheet[f'I{start_row_bs}'] =  row['total_bal11']
                bs_sheet[f'J{start_row_bs}'] =  row['total_bal12']
                bs_sheet[f'K{start_row_bs}'] =  row['total_bal1']
                bs_sheet[f'L{start_row_bs}'] =  row['total_bal2']
                bs_sheet[f'M{start_row_bs}'] =  row['total_bal3']
                bs_sheet[f'N{start_row_bs}'] =  row['total_bal4']
                bs_sheet[f'O{start_row_bs}'] =  row['total_bal5']
                bs_sheet[f'P{start_row_bs}'] =  row['total_bal6']
                bs_sheet[f'Q{start_row_bs}'] =  row['total_bal7']
                bs_sheet[f'R{start_row_bs}'] =  row['total_bal8']
                bs_sheet[f'T{start_row_bs}'] =  row['fytd']
                last_month_row_bal =f'total_bal{months["last_month_number"]}'
                bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

                                

        
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for col in range(5, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.style = currency_style_noborder 

        
        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Cash':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            for col in range(5, 22):  
                                cell = bs_sheet.cell(row=start_row, column=col)
                                cell.style = currency_style
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            print("324")
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            cash_row_bs = start_row_bs

                            for col in range(last_number,19):
                                col_letter = get_column_letter(col)
                                cell = bs_sheet.cell(row=start_row_bs, column=col)
                                cell.value = None


        hide_row_bs_end = None
        hide_row_bs_start = start_row_bs
        for row in data_activitybs:
            if row['Activity'] == 'Restr':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']

                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                try:
                    bs_sheet.row_dimensions[row].outline_level = 1
                    bs_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Restr':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            for col in range(5, 22):  # Columns G to U
                                cell = bs_sheet.cell(row=start_row_bs, column=col)
                                cell.style = normal_cell 
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            restr_row_bs = start_row_bs


        hide_row_bs_end = None
        hide_row_bs_start = start_row_bs
        for row in data_activitybs: 
            if row['Activity'] == 'DFS+F':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']

                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'DFS+F':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            dfs_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs 
        hide_row_bs_end = None  
        for row in data_activitybs: 
            if row['Activity'] == 'OTHR':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']

                    
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

                

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'OTHR':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            othr_row_bs = start_row_bs

        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'Inventory':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]
            
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Inventory':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            inventory_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'PPD':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'PPD':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            ppd_row_bs = start_row_bs



        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border    
        total_current_assets_row_bs = start_row_bs
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total Current Assets'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_assets_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_assets"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_assets"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_assets"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_assets"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_assets"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_assets"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_assets"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_assets"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_assets"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_assets"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_assets"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_assets"]["08"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_current_assets_fytd"]
        
        
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_current_assets"][acc_per]
        
        start_row_bs += 1
        bs_sheet[f'D{start_row_bs}'] = 'Capital Assets , Net'
        bs_sheet.row_dimensions[start_row_bs].height = 37 



        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'FA-L':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'FA-L':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Capital Assets, Net':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            fal_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs
        hide_row_bs_end = None   
        for row in data_activitybs: 
            if row['Activity'] == 'FA-BFE':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'FA-BFE':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Capital Assets, Net':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            fabfe_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'FA-AD':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'FA-AD':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Capital Assets, Net':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            faad_row_bs = start_row_bs


        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border    
        total_capital_assets_row_bs = start_row_bs
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total Capital Assets'
        
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_capital_assets_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_capital_assets"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_capital_assets"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_capital_assets"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_capital_assets"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_capital_assets"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_capital_assets"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_capital_assets"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_capital_assets"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_capital_assets"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_capital_assets"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_capital_assets"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_capital_assets"]["08"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_capital_assets_fytd"]
        
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_capital_assets"][acc_per]
        
        start_row_bs += 1
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold

        for col in range(6, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            
            cell.style = currency_style

        total_assets_row_bs = start_row_bs
    
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total  Assets'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_assets_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_assets"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_assets"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_assets"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_assets"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_assets"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_assets"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_assets"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_assets"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_assets"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_assets"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_assets"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_assets"]["08"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_assets_fye_fytd"]
     
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_assets"][acc_per]

        start_row_bs += 1
        bs_sheet.row_dimensions[start_row_bs].height = 37 
        
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Liabilities and Net Assets'
        
        start_row_bs += 1
        bs_sheet.row_dimensions[start_row_bs].height = 37 
        bs_sheet[f'D{start_row_bs}'] = 'Current Liabilities'


        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'AP':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]
    
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")      

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'AP':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            ap_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs 
        hide_row_bs_end = None  
        for row in data_activitybs: 
            if row['Activity'] == 'Acc-Exp':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Acc-Exp':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            accexp_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'OtherLiab':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]
            
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'OtherLiab':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            otherlab_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'Debt-C':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Debt-C':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            debtc_row_bs = start_row_bs

        hide_row_bs_start = start_row_bs 
        hide_row_bs_end = None  
        for row in data_activitybs: 
            if row['Activity'] == 'ACC-Int':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        
        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'ACC-Int':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            accint_row_bs = start_row_bs

        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border    

        total_current_liabilites_row_bs = start_row_bs
        
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
    
        bs_sheet[f'D{start_row_bs}'] ='Total Current Liabilities'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_liabilities_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_liabilities"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_liabilities"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_liabilities"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_liabilities"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_liabilities"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_liabilities"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_liabilities"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_liabilities"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_liabilities"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_liabilities"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_liabilities"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_liabilities"]["08"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_current_liabilities_fytd"]
       
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_current_liabilities"][acc_per]

        start_row_bs += 1
        bs_sheet.row_dimensions[start_row_bs].height = 37 
        bs_sheet[f'D{start_row_bs}'] ='Long Term Debt'


        hide_row_bs_start = start_row_bs
        hide_row_bs_end = None   
        for row in data_activitybs: 
            if row['Activity'] == 'LTD':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]



        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")



        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'LTD':
                    if row['Category'] == 'Debt':
                        if row['Subcategory'] == 'Long Term Debt':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            ltd_row_bs = start_row_bs

        start_row_bs += 1
        total_liabilites_row_bs = start_row_bs
    
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total Liabilities'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_liabilities_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_liabilities"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_liabilities"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_liabilities"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_liabilities"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_liabilities"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_liabilities"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_liabilities"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_liabilities"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_liabilities"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_liabilities"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_liabilities"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_liabilities"]["08"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_liabilities_fytd"]
        
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_liabilities"][acc_per]


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Equity':
                    if row['Category'] == 'Net Assets':
                    
                        start_row_bs += 1
                        
                        bs_sheet[f'D{start_row_bs}'].font = fontbold
                        bs_sheet[f'D{start_row_bs}'] = 'Net Assets'

                        bs_sheet[f'F{start_row_bs}'] = row['FYE']
                        bs_sheet[f'G{start_row_bs}'] = row['net_assets9']
                        bs_sheet[f'H{start_row_bs}'] = row['net_assets10']
                        bs_sheet[f'I{start_row_bs}'] = row['net_assets11']
                        bs_sheet[f'J{start_row_bs}'] = row['net_assets12']
                        bs_sheet[f'K{start_row_bs}'] = row['net_assets1']
                        bs_sheet[f'L{start_row_bs}'] = row['net_assets2']
                        bs_sheet[f'M{start_row_bs}'] = row['net_assets3']
                        bs_sheet[f'N{start_row_bs}'] = row['net_assets4']
                        bs_sheet[f'O{start_row_bs}'] = row['net_assets5']
                        bs_sheet[f'P{start_row_bs}'] = row['net_assets6']
                        bs_sheet[f'Q{start_row_bs}'] = row['net_assets7']
                        bs_sheet[f'R{start_row_bs}'] = row['net_assets8']
                        bs_sheet[f'T{start_row_bs}'] = total_bs['total_net_assets_fytd']
                        last_month_row_net =f'net_assets{months["last_month_number"]}'
                        bs_sheet[f'U{start_row_bs}'] = row[last_month_row_net]
                        net_assets_row_bs = start_row_bs


        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border
            cell.style = currency_style    
        
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
    
        bs_sheet[f'D{start_row_bs}'] = 'Total Liabilities and Net Assets'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_LNA_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_LNA"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_LNA"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_LNA"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_LNA"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_LNA"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_LNA"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_LNA"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_LNA"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_LNA"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_LNA"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_LNA"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_LNA"]["08"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_LNA_fytd"]
      
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_LNA"][acc_per]
    


        while start_bs_for_hiding <= start_row_bs:
            for col in range(last_number,19):
                col_letter = get_column_letter(col)
                cell = bs_sheet.cell(row=start_bs_for_hiding, column=col)
                cell.value = None
            start_bs_for_hiding += 1   


    else:
        for col in range(7, 20 ):
            col_letter = get_column_letter(col)
            bs_sheet.column_dimensions[col_letter].outline_level = 1
            bs_sheet.column_dimensions[col_letter].hidden = True

        last_number = months["last_month_number"]
        # PL START OF DESIGN
        if last_number <= 6:
            last_number += 13
        else:
            last_number += 1

        acc_per =  months["last_month_number"]

        if acc_per >= 10:
            acc_per = str(acc_per)
            print("eto")
        else:
            acc_per = f'0{acc_per}'

        for col in range(last_number,19):
            col_letter = get_column_letter(col)     
            bs_sheet.column_dimensions[col_letter].outline_level = 2
            bs_sheet.column_dimensions[col_letter].hidden = True


        header_bs = 3
        bs_sheet[f'G{header_bs}'] = 'July'

        bs_sheet[f'H{header_bs}'] = 'August'

        bs_sheet[f'I{header_bs}'] = 'September'

        bs_sheet[f'J{header_bs}'] = 'October'

        bs_sheet[f'K{header_bs}'] = 'November'

        bs_sheet[f'L{header_bs}'] = 'December'

        bs_sheet[f'M{header_bs}'] = 'January'

        bs_sheet[f'N{header_bs}'] = 'February'

        bs_sheet[f'O{header_bs}'] = 'March'

        bs_sheet[f'P{header_bs}'] = 'April'

        bs_sheet[f'Q{header_bs}'] = 'May'

        bs_sheet[f'R{header_bs}'] = 'June'



       
        start_bs_for_hiding  = 7  
        start_row_bs = 6
        hide_row_bs_start = start_row_bs
        hide_row_bs_end = None
        bs_sheet[f'D{start_row_bs}'] = 'Current Assets'
        for row in data_activitybs:
            if row['Activity'] == 'Cash':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22):  # Columns G to U
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] =  row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] =  row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] =  row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] =  row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] =  row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] =  row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] =  row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] =  row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] =  row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] =  row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] =  row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] =  row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] =  row['fytd']

                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for col in range(5, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.style = currency_style_noborder 

        
        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Cash':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            for col in range(5, 22):  
                                cell = pl_sheet.cell(row=start_row, column=col)
                                cell.style = currency_style
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            cash_row_bs = start_row_bs


        
        hide_row_bs_start = start_row_bs
        hide_row_bs_end = None
        for row in data_activitybs:
            if row['Activity'] == 'Restr':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']

                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                try:
                    bs_sheet.row_dimensions[row].outline_level = 1
                    bs_sheet.row_dimensions[row].hidden = True
                except KeyError as e:
                    print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Restr':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            for col in range(5, 22):  # Columns G to U
                                cell = bs_sheet.cell(row=start_row_bs, column=col)
                                cell.style = normal_cell 
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            restr_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'DFS+F':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']

                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'DFS+F':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            dfs_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs 
        hide_row_bs_end = None  
        for row in data_activitybs: 
            if row['Activity'] == 'OTHR':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']

                    
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

                

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'OTHR':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']

                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            othr_row_bs = start_row_bs

        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'Inventory':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]
            
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Inventory':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            inventory_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs
        hide_row_bs_end = None   
        for row in data_activitybs: 
            if row['Activity'] == 'PPD':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'PPD':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Current Assets':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            ppd_row_bs = start_row_bs



        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border    
        total_current_assets_row_bs = start_row_bs
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total Current Assets'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_assets_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_assets"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_assets"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_assets"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_assets"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_assets"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_assets"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_assets"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_assets"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_assets"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_assets"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_assets"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_assets"]["06"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_current_assets_fytd"]
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_current_assets"][acc_per]
        
        start_row_bs += 1
        bs_sheet[f'D{start_row_bs}'] = 'Capital Assets , Net'
        bs_sheet.row_dimensions[start_row_bs].height = 37 



        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'FA-L':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros: 
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'FA-L':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Capital Assets, Net':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            fal_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'FA-BFE':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'FA-BFE':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Capital Assets, Net':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            fabfe_row_bs = start_row_bs



        hide_row_bs_start = start_row_bs 
        hide_row_bs_end = None  
        for row in data_activitybs: 
            if row['Activity'] == 'FA-AD':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'FA-AD':
                    if row['Category'] == 'Assets':
                        if row['Subcategory'] == 'Capital Assets, Net':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                            bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                            bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                            bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                            bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                            bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                            bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                            bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                            bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                            bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                            bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                            bs_sheet[f'T{start_row_bs}'] = row['fytd']
                            last_month_row = f'difference_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            faad_row_bs = start_row_bs


        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border    
        total_capital_assets_row_bs = start_row_bs
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total Capital Assets'
        
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_capital_assets_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_capital_assets"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_capital_assets"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_capital_assets"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_capital_assets"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_capital_assets"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_capital_assets"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_capital_assets"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_capital_assets"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_capital_assets"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_capital_assets"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_capital_assets"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_capital_assets"]["06"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_capital_assets_fytd"]
 
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_capital_assets"][acc_per]
        
        start_row_bs += 1
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold

        for col in range(6, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            
            cell.style = currency_style

        total_assets_row_bs = start_row_bs
    
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total  Assets'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_assets_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_assets"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_assets"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_assets"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_assets"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_assets"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_assets"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_assets"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_assets"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_assets"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_assets"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_assets"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_assets"]["06"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_assets_fye_fytd"]
      
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_assets"][acc_per]

        start_row_bs += 1
        bs_sheet.row_dimensions[start_row_bs].height = 37 
        
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Liabilities and Net Assets'
        
        start_row_bs += 1
        bs_sheet.row_dimensions[start_row_bs].height = 37 
        bs_sheet[f'D{start_row_bs}'] = 'Current Liabilities'


        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'AP':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]
    

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")      

        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'AP':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            ap_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'Acc-Exp':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Acc-Exp':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            accexp_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'OtherLiab':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]
                
        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'OtherLiab':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            otherlab_row_bs = start_row_bs


        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'Debt-C':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Debt-C':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            debtc_row_bs = start_row_bs

        hide_row_bs_start = start_row_bs  
        hide_row_bs_end = None 
        for row in data_activitybs: 
            if row['Activity'] == 'ACC-Int':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]

        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")


        
        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'ACC-Int':
                    if row['Category'] == 'Liabilities and Net Assets':
                        if row['Subcategory'] == 'Current Liabilities':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            accint_row_bs = start_row_bs

        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border    

        total_current_liabilites_row_bs = start_row_bs
        
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
    
        bs_sheet[f'D{start_row_bs}'] ='Total Current Liabilities'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_liabilities_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_liabilities"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_liabilities"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_liabilities"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_liabilities"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_liabilities"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_liabilities"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_liabilities"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_liabilities"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_liabilities"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_liabilities"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_liabilities"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_liabilities"]["06"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_current_liabilities_fytd"]
     
        bs_sheet[f'U{start_row_bs}'] = total_bs["total_current_liabilities"][acc_per]

        start_row_bs += 1
        bs_sheet.row_dimensions[start_row_bs].height = 37 
        bs_sheet[f'D{start_row_bs}'] ='Long Term Debt'


        hide_row_bs_start = start_row_bs   
        hide_row_bs_end = None
        for row in data_activitybs: 
            if row['Activity'] == 'LTD':
                all_zeros = all(row[f'total_bal{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    start_row_bs += 1
                    hide_row_bs_end = start_row_bs
                    for col in range(5, 22): 
                        cell = bs_sheet.cell(row=start_row_bs, column=col)
                        cell.style = normal_cell 
                    bs_sheet[f'D{start_row_bs}'].style = indent_style
                    bs_sheet[f'D{start_row_bs}'] = row['obj'] + ' - ' + row['Description2']
                    bs_sheet[f'G{start_row_bs}'] = row['total_bal7']
                    bs_sheet[f'H{start_row_bs}'] = row['total_bal8']
                    bs_sheet[f'I{start_row_bs}'] = row['total_bal9']
                    bs_sheet[f'J{start_row_bs}'] = row['total_bal10']
                    bs_sheet[f'K{start_row_bs}'] = row['total_bal11']
                    bs_sheet[f'L{start_row_bs}'] = row['total_bal12']
                    bs_sheet[f'M{start_row_bs}'] = row['total_bal1']
                    bs_sheet[f'N{start_row_bs}'] = row['total_bal2']
                    bs_sheet[f'O{start_row_bs}'] = row['total_bal3']
                    bs_sheet[f'P{start_row_bs}'] = row['total_bal4']
                    bs_sheet[f'Q{start_row_bs}'] = row['total_bal5']
                    bs_sheet[f'R{start_row_bs}'] = row['total_bal6']
                    bs_sheet[f'T{start_row_bs}'] = row['fytd']
                    last_month_row_bal =f'total_bal{months["last_month_number"]}'
                    bs_sheet[f'U{start_row_bs}'] = row[last_month_row_bal]


        if hide_row_bs_end is not None:
            for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                    try:
                        bs_sheet.row_dimensions[row].outline_level = 1
                        bs_sheet.row_dimensions[row].hidden = True

                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")



        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'LTD':
                    if row['Category'] == 'Debt':
                        if row['Subcategory'] == 'Long Term Debt':
                            start_row_bs += 1
                            bs_sheet[f'D{start_row_bs}'].style = indent_style
                            bs_sheet[f'D{start_row_bs}'] = row['Description']
                            bs_sheet[f'F{start_row_bs}'] = row['FYE']
                            bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                            bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                            bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                            bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                            bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                            bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                            bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                            bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                            bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                            bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                            bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                            bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                            bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                            last_month_row = f'debt_{months["last_month_number"]}'
                            bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                            ltd_row_bs = start_row_bs

        start_row_bs += 1
        total_liabilites_row_bs = start_row_bs
    
        bs_sheet[f'D{start_row_bs}'].font = fontbold
        bs_sheet[f'D{start_row_bs}'] = 'Total Liabilities'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_liabilities_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_liabilities"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_liabilities"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_liabilities"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_liabilities"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_liabilities"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_liabilities"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_liabilities"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_liabilities"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_liabilities"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_liabilities"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_liabilities"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_liabilities"]["06"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_liabilities_fytd"]

        bs_sheet[f'U{start_row_bs}'] = total_bs["total_liabilities"][acc_per]


        for row in data_balancesheet:
            if row['school'] == school:
                if row['Activity'] == 'Equity':
                    if row['Category'] == 'Net Assets':
                    
                        start_row_bs += 1
                        
                        bs_sheet[f'D{start_row_bs}'].font = fontbold
                        bs_sheet[f'D{start_row_bs}'] = 'Net Assets'

                        bs_sheet[f'F{start_row_bs}'] = row['FYE']
                        bs_sheet[f'G{start_row_bs}'] = row['net_assets7']
                        bs_sheet[f'H{start_row_bs}'] = row['net_assets8']
                        bs_sheet[f'I{start_row_bs}'] = row['net_assets9']
                        bs_sheet[f'J{start_row_bs}'] = row['net_assets10']
                        bs_sheet[f'K{start_row_bs}'] = row['net_assets11']
                        bs_sheet[f'L{start_row_bs}'] = row['net_assets12']
                        bs_sheet[f'M{start_row_bs}'] = row['net_assets1']
                        bs_sheet[f'N{start_row_bs}'] = row['net_assets2']
                        bs_sheet[f'O{start_row_bs}'] = row['net_assets3']
                        bs_sheet[f'P{start_row_bs}'] = row['net_assets4']
                        bs_sheet[f'Q{start_row_bs}'] = row['net_assets5']
                        bs_sheet[f'R{start_row_bs}'] = row['net_assets6']
                        bs_sheet[f'T{start_row_bs}'] = total_bs['total_net_assets_fytd']
                        last_month_row_net =f'net_assets{months["last_month_number"]}'
                        bs_sheet[f'U{start_row_bs}'] = row[last_month_row_net]
                        net_assets_row_bs = start_row_bs


        start_row_bs += 1    
        for col in range(2, 22):  
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.font = fontbold
        for col in range(6, 22):  # Columns D to U
            cell = bs_sheet.cell(row=start_row_bs, column=col)
            cell.border = thin_border
            cell.style = currency_style    
        
        bs_sheet[f'D{start_row_bs}'].style = indent_style2
        bs_sheet[f'D{start_row_bs}'].font = fontbold
    
        bs_sheet[f'D{start_row_bs}'] = 'Total Liabilities and Net Assets'
        bs_sheet[f'F{start_row_bs}'] = total_bs["total_LNA_fye"]
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_LNA"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_LNA"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_LNA"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_LNA"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_LNA"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_LNA"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_LNA"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_LNA"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_LNA"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_LNA"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_LNA"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_LNA"]["06"]
        bs_sheet[f'T{start_row_bs}'] = total_bs["total_LNA_fytd"]

        bs_sheet[f'U{start_row_bs}'] = total_bs["total_LNA"][acc_per]
        
        while start_bs_for_hiding <= start_row_bs:
            for col in range(last_number,19):
                col_letter = get_column_letter(col)
                cell = bs_sheet.cell(row=start_bs_for_hiding, column=col)
                cell.value = None
            start_bs_for_hiding += 1 



    #CASHFLOW DESIGN
    for row in range(2,181):
        cashflow_sheet.row_dimensions[row].height = 19
    # cashflow_sheet.row_dimensions[17].height = 26 #local revenue
    # cashflow_sheet.row_dimensions[20].height = 26 #spr
    # cashflow_sheet.row_dimensions[33].height = 26 #fpr
    # cashflow_sheet.row_dimensions[34].height = 26 
    cashflow_sheet.column_dimensions['A'].width = 5
    cashflow_sheet.column_dimensions['B'].width = 46
    cashflow_sheet.column_dimensions['C'].width = 10
    cashflow_sheet.column_dimensions['D'].width = 14
    cashflow_sheet.column_dimensions['E'].width = 14
    cashflow_sheet.column_dimensions['F'].hidden = True
    cashflow_sheet.column_dimensions['G'].width = 14
    cashflow_sheet.column_dimensions['H'].width = 14
    cashflow_sheet.column_dimensions['I'].width = 14
    cashflow_sheet.column_dimensions['J'].width = 14
    cashflow_sheet.column_dimensions['K'].width = 14
    cashflow_sheet.column_dimensions['L'].width = 14
    cashflow_sheet.column_dimensions['M'].width = 14
    cashflow_sheet.column_dimensions['N'].width = 14
    cashflow_sheet.column_dimensions['O'].width = 14
    cashflow_sheet.column_dimensions['P'].width = 3
    cashflow_sheet.column_dimensions['Q'].width = 14
    cashflow_sheet.column_dimensions['R'].width = 14
    cashflow_sheet.column_dimensions['S'].width = 3
    cashflow_sheet.column_dimensions['T'].width = 17
    cashflow_sheet.column_dimensions['U'].width = 17
    cashflow_sheet.column_dimensions['V'].width = 14




    start = 1 
    cashflow_sheet[f'A{start}'] = school_name
    start += 1
    cashflow_sheet[f'A{start}'] = 'Statement of Cash Flows'
    start += 1
    cashflow_sheet[f'A{start}'] = f'for the period ended of {months["last_month"]}'

    thin_border = Border(top=Side(border_style='thin'))
    currency_style = NamedStyle(name="currency_style", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style.border =Border(top=Side(border_style='thin'))
    currency_style.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style.font = currency_font


    normal_cell_bottom_border = NamedStyle(name="normal_cell_bottom_border", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell_bottom_border.border =Border(bottom=Side(border_style='thin'))
    normal_cell_bottom_border.alignment = Alignment(horizontal='right', vertical='bottom')
    normal_font_bottom_border = Font(name='Calibri', size=11, bold=False)
    normal_cell_bottom_border.font = normal_font_bottom_border

    normal_cell = NamedStyle(name="normal_cell", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell.alignment = Alignment(horizontal='right', vertical='bottom')
    normal_font = Font(name='Calibri', size=11, bold=False)
    normal_cell.font = normal_font


    currency_style_noborder = NamedStyle(name="currency_style_noborder", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    
    currency_style_noborder.alignment = Alignment(horizontal='right', vertical='top')
    currency_font_noborder = Font(name='Calibri', size=11, bold=True)
    currency_style_noborder.font = currency_font_noborder
          

    thin_border_topdown = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    currency_style_topdown_border = NamedStyle(name="currency_style_topdown_border", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style_topdown_border.border = thin_border_topdown
    currency_style_topdown_border.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style_topdown_border.font = currency_font

    
    if  school != 'manara' and school != 'leadership':

        for col in range(4, 17):
            col_letter = get_column_letter(col)
            cashflow_sheet.column_dimensions[col_letter].outline_level = 1
            cashflow_sheet.column_dimensions[col_letter].hidden = True
        last_number = months["last_month_number"]
        
        # PL START OF DESIGN
        if last_number <= 8:
            last_number += 11
        else:
            last_number -= 1


        for col in range(last_number-3,16):
            col_letter = get_column_letter(col)
          
      
            cashflow_sheet.column_dimensions[col_letter].outline_level = 2
            cashflow_sheet.column_dimensions[col_letter].hidden = True

        

       

        cashflow_start_hiding = 7
        cashflow_start_row = 7
        
        operating_start_row = cashflow_start_row
        
        cashflow_sheet[f'B{cashflow_start_row}'] = 'Change in Net Assets'
        cashflow_sheet[f'D{cashflow_start_row}'] = totals["total_netsurplus_months"]["09"]
        cashflow_sheet[f'E{cashflow_start_row}'] = totals["total_netsurplus_months"]["10"]
        cashflow_sheet[f'F{cashflow_start_row}'] = totals["total_netsurplus_months"]["11"]
        cashflow_sheet[f'G{cashflow_start_row}'] = totals["total_netsurplus_months"]["12"]
        cashflow_sheet[f'H{cashflow_start_row}'] = totals["total_netsurplus_months"]["01"]
        cashflow_sheet[f'I{cashflow_start_row}'] = totals["total_netsurplus_months"]["02"]
        cashflow_sheet[f'J{cashflow_start_row}'] = totals["total_netsurplus_months"]["03"]
        cashflow_sheet[f'K{cashflow_start_row}'] = totals["total_netsurplus_months"]["04"]
        cashflow_sheet[f'L{cashflow_start_row}'] = totals["total_netsurplus_months"]["05"]
        cashflow_sheet[f'M{cashflow_start_row}'] = totals["total_netsurplus_months"]["06"]
        cashflow_sheet[f'N{cashflow_start_row}'] = totals["total_netsurplus_months"]["07"]
        cashflow_sheet[f'O{cashflow_start_row}'] = totals["total_netsurplus_months"]["08"]
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 
    
        cashflow_start_row += 2
        cashflow_sheet[f'B{cashflow_start_row}'] = 'Depreciation and Amortization'
        cashflow_sheet[f'D{cashflow_start_row}'] = totals["dna_total_months"]["09"]
        cashflow_sheet[f'E{cashflow_start_row}'] = totals["dna_total_months"]["10"]
        cashflow_sheet[f'F{cashflow_start_row}'] = totals["dna_total_months"]["11"]
        cashflow_sheet[f'G{cashflow_start_row}'] = totals["dna_total_months"]["12"]
        cashflow_sheet[f'H{cashflow_start_row}'] = totals["dna_total_months"]["01"]
        cashflow_sheet[f'I{cashflow_start_row}'] = totals["dna_total_months"]["02"]
        cashflow_sheet[f'J{cashflow_start_row}'] = totals["dna_total_months"]["03"]
        cashflow_sheet[f'K{cashflow_start_row}'] = totals["dna_total_months"]["04"]
        cashflow_sheet[f'L{cashflow_start_row}'] = totals["dna_total_months"]["05"]
        cashflow_sheet[f'M{cashflow_start_row}'] = totals["dna_total_months"]["06"]
        cashflow_sheet[f'N{cashflow_start_row}'] = totals["dna_total_months"]["07"]
        cashflow_sheet[f'O{cashflow_start_row}'] = totals["dna_total_months"]["08"]
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 

    
        #CASHFLOW FROM OPERATING ACTIVITIES
        for row in data_cashflow:
            if row['Category'] == 'Operating':
                all_zeros = all(row[f'total_operating{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    cashflow_start_row += 1
                    cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_operating9']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_operating10']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_operating11']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_operating12']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_operating1']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_operating2']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_operating3']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_operating4']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_operating5']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_operating6']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_operating7']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_operating8']
                    cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 

        operating_end_row = cashflow_start_row
        cashflow_start_row += 1
        net_operating_total_row = cashflow_start_row
        
        # NET OPERATING TOTAL
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        cashflow_sheet[f'A{cashflow_start_row}'] = 'NET CASH FLOWS FROM OPERATING ACTIVITIES'
        cashflow_sheet[f'D{cashflow_start_row}'].value = f'=SUM(D{operating_start_row}:D{operating_end_row})'  
        cashflow_sheet[f'E{cashflow_start_row}'].value = f'=SUM(E{operating_start_row}:E{operating_end_row})'
        cashflow_sheet[f'F{cashflow_start_row}'].value = f'=SUM(F{operating_start_row}:F{operating_end_row})' 
        cashflow_sheet[f'G{cashflow_start_row}'].value = f'=SUM(G{operating_start_row}:G{operating_end_row})' 
        cashflow_sheet[f'H{cashflow_start_row}'].value = f'=SUM(H{operating_start_row}:H{operating_end_row})' 
        cashflow_sheet[f'I{cashflow_start_row}'].value = f'=SUM(I{operating_start_row}:I{operating_end_row})' 
        cashflow_sheet[f'J{cashflow_start_row}'].value = f'=SUM(J{operating_start_row}:J{operating_end_row})' 
        cashflow_sheet[f'K{cashflow_start_row}'].value = f'=SUM(K{operating_start_row}:K{operating_end_row})' 
        cashflow_sheet[f'L{cashflow_start_row}'].value = f'=SUM(L{operating_start_row}:L{operating_end_row})' 
        cashflow_sheet[f'M{cashflow_start_row}'].value = f'=SUM(M{operating_start_row}:M{operating_end_row})' 
        cashflow_sheet[f'N{cashflow_start_row}'].value = f'=SUM(N{operating_start_row}:N{operating_end_row})' 
        cashflow_sheet[f'O{cashflow_start_row}'].value = f'=SUM(O{operating_start_row}:O{operating_end_row})' 
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(Q{operating_start_row}:Q{operating_end_row})' 

        

        cashflow_start_row += 1

        investing_row_start = cashflow_start_row
        #CASHFLOW FROM INVESTING ACTIVITIES
        for row in data_cashflow:
            if row['Category'] == 'Investing':
                all_zeros = all(row[f'total_investing{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    cashflow_start_row += 1
                    
                    cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
            
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_investing9']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_investing10']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_investing11']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_investing12']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_investing1']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_investing2']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_investing3']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_investing4']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_investing5']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_investing6']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_investing7']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_investing8']
                    cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 

        investing_row_end = cashflow_start_row
        cashflow_start_row += 1
        #NET INVESTING TOTAL
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}")
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 

        net_investing_total_row = cashflow_start_row
        cashflow_sheet[f'A{cashflow_start_row}'] = 'NET CASH FLOWS FROM INVESTING ACTIVITIES'
        cashflow_sheet[f'D{cashflow_start_row}'].value = f'=SUM(D{investing_row_start}:D{investing_row_end})'  
        cashflow_sheet[f'E{cashflow_start_row}'].value = f'=SUM(E{investing_row_start}:E{investing_row_end})'
        cashflow_sheet[f'F{cashflow_start_row}'].value = f'=SUM(F{investing_row_start}:F{investing_row_end})' 
        cashflow_sheet[f'G{cashflow_start_row}'].value = f'=SUM(G{investing_row_start}:G{investing_row_end})' 
        cashflow_sheet[f'H{cashflow_start_row}'].value = f'=SUM(H{investing_row_start}:H{investing_row_end})' 
        cashflow_sheet[f'I{cashflow_start_row}'].value = f'=SUM(I{investing_row_start}:I{investing_row_end})' 
        cashflow_sheet[f'J{cashflow_start_row}'].value = f'=SUM(J{investing_row_start}:J{investing_row_end})' 
        cashflow_sheet[f'K{cashflow_start_row}'].value = f'=SUM(K{investing_row_start}:K{investing_row_end})' 
        cashflow_sheet[f'L{cashflow_start_row}'].value = f'=SUM(L{investing_row_start}:L{investing_row_end})' 
        cashflow_sheet[f'M{cashflow_start_row}'].value = f'=SUM(M{investing_row_start}:M{investing_row_end})' 
        cashflow_sheet[f'N{cashflow_start_row}'].value = f'=SUM(N{investing_row_start}:N{investing_row_end})' 
        cashflow_sheet[f'O{cashflow_start_row}'].value = f'=SUM(O{investing_row_start}:O{investing_row_end})' 
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(Q{investing_row_start}:Q{investing_row_end})' 

    
        #NET INCREASE Decrease in cash
        cashflow_start_row += 2
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        cashflow_sheet[f'A{cashflow_start_row}'] = 'NET INCREASE (DECREASE IN CASH)'
        cashflow_sheet[f'D{cashflow_start_row}'].value = f'=SUM(D{net_operating_total_row},D{net_investing_total_row})'  
        cashflow_sheet[f'E{cashflow_start_row}'].value = f'=SUM(E{net_operating_total_row},E{net_investing_total_row})'
        cashflow_sheet[f'F{cashflow_start_row}'].value = f'=SUM(F{net_operating_total_row},F{net_investing_total_row})' 
        cashflow_sheet[f'G{cashflow_start_row}'].value = f'=SUM(G{net_operating_total_row},G{net_investing_total_row})' 
        cashflow_sheet[f'H{cashflow_start_row}'].value = f'=SUM(H{net_operating_total_row},H{net_investing_total_row})' 
        cashflow_sheet[f'I{cashflow_start_row}'].value = f'=SUM(I{net_operating_total_row},I{net_investing_total_row})' 
        cashflow_sheet[f'J{cashflow_start_row}'].value = f'=SUM(J{net_operating_total_row},J{net_investing_total_row})' 
        cashflow_sheet[f'K{cashflow_start_row}'].value = f'=SUM(K{net_operating_total_row},K{net_investing_total_row})' 
        cashflow_sheet[f'L{cashflow_start_row}'].value = f'=SUM(L{net_operating_total_row},L{net_investing_total_row})' 
        cashflow_sheet[f'M{cashflow_start_row}'].value = f'=SUM(M{net_operating_total_row},M{net_investing_total_row})' 
        cashflow_sheet[f'N{cashflow_start_row}'].value = f'=SUM(N{net_operating_total_row},N{net_investing_total_row})' 
        cashflow_sheet[f'O{cashflow_start_row}'].value = f'=SUM(O{net_operating_total_row},O{net_investing_total_row})' 
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(Q{net_operating_total_row},Q{net_investing_total_row})' 


        cashflow_start_row  += 2
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        for row in data_balancesheet:
            if row["school"] == school:
                if row['Category'] == 'Assets':
                    if row['Subcategory'] == 'Current Assets':
                        if row['Activity'] == 'Cash':
                            cashflow_sheet[f'A{cashflow_start_row}'] = 'Cash At Beginning of Period'
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['FYE']
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_10']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_6']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'Q{cashflow_start_row}'] = row['FYE']

        cashflow_start_row  += 2
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}")

        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 

        for row in data_balancesheet:
            if row["school"] == school:
                if row['Category'] == 'Assets':
                    if row['Subcategory'] == 'Current Assets':
                        if row['Activity'] == 'Cash':
                            cashflow_sheet[f'A{cashflow_start_row}'] = 'Cash At End of Period'
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_6']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_8']
                            cashflow_sheet[f'Q{cashflow_start_row}'] = row['last_month_difference']

        while cashflow_start_hiding <= cashflow_start_row:
            for col in range(last_number-3,16):
                col_letter = get_column_letter(col)
             
                cell = cashflow_sheet.cell(row=cashflow_start_hiding, column=col)
                cell.value = None
            cashflow_start_hiding += 1


    else:
        #MANARA CASHFLOW
        header_cashflow = 5
        bs_sheet[f'D{header_cashflow}'] = 'July'
        bs_sheet[f'E{header_cashflow}'] = 'August'
        bs_sheet[f'F{header_cashflow}'] = 'September'
        bs_sheet[f'G{header_cashflow}'] = 'October'
        bs_sheet[f'H{header_cashflow}'] = 'November'
        bs_sheet[f'I{header_cashflow}'] = 'December'
        bs_sheet[f'J{header_cashflow}'] = 'January'
        bs_sheet[f'K{header_cashflow}'] = 'February'
        bs_sheet[f'L{header_cashflow}'] = 'March'
        bs_sheet[f'M{header_cashflow}'] = 'April'
        bs_sheet[f'N{header_cashflow}'] = 'May'
        bs_sheet[f'O{header_cashflow}'] = 'June'
        last_number = months["last_month_number"]
        # PL START OF DESIGN


  

        for col in range(4, 17):
            col_letter = get_column_letter(col)
            cashflow_sheet.column_dimensions[col_letter].outline_level = 1
            cashflow_sheet.column_dimensions[col_letter].hidden = True
        last_number = months["last_month_number"]
        
     
        if last_number <= 6:
            last_number += 13
        else:
            last_number += 1


        for col in range(last_number-3,16):
            col_letter = get_column_letter(col)
          
      
            cashflow_sheet.column_dimensions[col_letter].outline_level = 2
            cashflow_sheet.column_dimensions[col_letter].hidden = True

        
        cashflow_start_row = 7
        cashflow_start_hiding = 7
        operating_start_row = cashflow_start_row
        cashflow_sheet[f'B{cashflow_start_row}'] = 'Change in Net Assets'
        cashflow_sheet[f'D{cashflow_start_row}'] = totals["total_netsurplus_months"]["07"]
        cashflow_sheet[f'E{cashflow_start_row}'] = totals["total_netsurplus_months"]["08"]
        cashflow_sheet[f'F{cashflow_start_row}'] = totals["total_netsurplus_months"]["09"]
        cashflow_sheet[f'G{cashflow_start_row}'] = totals["total_netsurplus_months"]["10"]
        cashflow_sheet[f'H{cashflow_start_row}'] = totals["total_netsurplus_months"]["11"]
        cashflow_sheet[f'I{cashflow_start_row}'] = totals["total_netsurplus_months"]["12"]
        cashflow_sheet[f'J{cashflow_start_row}'] = totals["total_netsurplus_months"]["01"]
        cashflow_sheet[f'K{cashflow_start_row}'] = totals["total_netsurplus_months"]["02"]
        cashflow_sheet[f'L{cashflow_start_row}'] = totals["total_netsurplus_months"]["03"]
        cashflow_sheet[f'M{cashflow_start_row}'] = totals["total_netsurplus_months"]["04"]
        cashflow_sheet[f'N{cashflow_start_row}'] = totals["total_netsurplus_months"]["05"]
        cashflow_sheet[f'O{cashflow_start_row}'] = totals["total_netsurplus_months"]["06"]
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 
    
        cashflow_start_row += 2
        cashflow_sheet[f'B{cashflow_start_row}'] = 'Depreciation and Amortization'
        cashflow_sheet[f'D{cashflow_start_row}'] = totals["dna_total_months"]["07"]
        cashflow_sheet[f'E{cashflow_start_row}'] = totals["dna_total_months"]["08"]
        cashflow_sheet[f'F{cashflow_start_row}'] = totals["dna_total_months"]["09"]
        cashflow_sheet[f'G{cashflow_start_row}'] = totals["dna_total_months"]["10"]
        cashflow_sheet[f'H{cashflow_start_row}'] = totals["dna_total_months"]["11"]
        cashflow_sheet[f'I{cashflow_start_row}'] = totals["dna_total_months"]["12"]
        cashflow_sheet[f'J{cashflow_start_row}'] = totals["dna_total_months"]["01"]
        cashflow_sheet[f'K{cashflow_start_row}'] = totals["dna_total_months"]["02"]
        cashflow_sheet[f'L{cashflow_start_row}'] = totals["dna_total_months"]["03"]
        cashflow_sheet[f'M{cashflow_start_row}'] = totals["dna_total_months"]["04"]
        cashflow_sheet[f'N{cashflow_start_row}'] = totals["dna_total_months"]["05"]
        cashflow_sheet[f'O{cashflow_start_row}'] = totals["dna_total_months"]["06"]
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 

    
        #CASHFLOW FROM OPERATING ACTIVITIES
        for row in data_cashflow:
            if row['Category'] == 'Operating':
                all_zeros = all(row[f'total_operating{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    cashflow_start_row += 1
                    cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_operating7']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_operating8']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_operating9']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_operating10']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_operating11']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_operating12']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_operating1']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_operating2']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_operating3']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_operating4']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_operating5']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_operating6']
                    cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 

        operating_end_row = cashflow_start_row
        cashflow_start_row = 19
        net_operating_total_row = cashflow_start_row
        
        # NET OPERATING TOTAL
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        cashflow_sheet[f'A{cashflow_start_row}'] = 'NET CASH FLOWS FROM OPERATING ACTIVITIES'
        cashflow_sheet[f'D{cashflow_start_row}'].value = f'=SUM(D{operating_start_row}:D{operating_end_row})'  
        cashflow_sheet[f'E{cashflow_start_row}'].value = f'=SUM(E{operating_start_row}:E{operating_end_row})'
        cashflow_sheet[f'F{cashflow_start_row}'].value = f'=SUM(F{operating_start_row}:F{operating_end_row})' 
        cashflow_sheet[f'G{cashflow_start_row}'].value = f'=SUM(G{operating_start_row}:G{operating_end_row})' 
        cashflow_sheet[f'H{cashflow_start_row}'].value = f'=SUM(H{operating_start_row}:H{operating_end_row})' 
        cashflow_sheet[f'I{cashflow_start_row}'].value = f'=SUM(I{operating_start_row}:I{operating_end_row})' 
        cashflow_sheet[f'J{cashflow_start_row}'].value = f'=SUM(J{operating_start_row}:J{operating_end_row})' 
        cashflow_sheet[f'K{cashflow_start_row}'].value = f'=SUM(K{operating_start_row}:K{operating_end_row})' 
        cashflow_sheet[f'L{cashflow_start_row}'].value = f'=SUM(L{operating_start_row}:L{operating_end_row})' 
        cashflow_sheet[f'M{cashflow_start_row}'].value = f'=SUM(M{operating_start_row}:M{operating_end_row})' 
        cashflow_sheet[f'N{cashflow_start_row}'].value = f'=SUM(N{operating_start_row}:N{operating_end_row})' 
        cashflow_sheet[f'O{cashflow_start_row}'].value = f'=SUM(O{operating_start_row}:O{operating_end_row})' 
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(Q{operating_start_row}:Q{operating_end_row})' 


        cashflow_start_row = 23

        investing_row_start = cashflow_start_row
        #CASHFLOW FROM INVESTING ACTIVITIES
        for row in data_cashflow:
            if row['Category'] == 'Investing':
                all_zeros = all(row[f'total_investing{i}'] == 0 for i in range(1, 12))
                if not all_zeros:
                    cashflow_start_row += 1
                    
                    cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
            
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_investing7']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_investing8']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_investing9']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_investing10']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_investing11']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_investing12']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_investing1']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_investing2']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_investing3']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_investing4']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_investing5']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_investing6']
                    cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(D{cashflow_start_row}:O{cashflow_start_row})' 

        investing_row_end = cashflow_start_row


        cashflow_start_row +1
        
        #NET INVESTING TOTAL
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}")
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 

        net_investing_total_row = cashflow_start_row
        cashflow_sheet[f'A{cashflow_start_row}'] = 'NET CASH FLOWS FROM INVESTING ACTIVITIES'
        cashflow_sheet[f'D{cashflow_start_row}'].value = f'=SUM(D{investing_row_start}:D{investing_row_end})'  
        cashflow_sheet[f'E{cashflow_start_row}'].value = f'=SUM(E{investing_row_start}:E{investing_row_end})'
        cashflow_sheet[f'F{cashflow_start_row}'].value = f'=SUM(F{investing_row_start}:F{investing_row_end})' 
        cashflow_sheet[f'G{cashflow_start_row}'].value = f'=SUM(G{investing_row_start}:G{investing_row_end})' 
        cashflow_sheet[f'H{cashflow_start_row}'].value = f'=SUM(H{investing_row_start}:H{investing_row_end})' 
        cashflow_sheet[f'I{cashflow_start_row}'].value = f'=SUM(I{investing_row_start}:I{investing_row_end})' 
        cashflow_sheet[f'J{cashflow_start_row}'].value = f'=SUM(J{investing_row_start}:J{investing_row_end})' 
        cashflow_sheet[f'K{cashflow_start_row}'].value = f'=SUM(K{investing_row_start}:K{investing_row_end})' 
        cashflow_sheet[f'L{cashflow_start_row}'].value = f'=SUM(L{investing_row_start}:L{investing_row_end})' 
        cashflow_sheet[f'M{cashflow_start_row}'].value = f'=SUM(M{investing_row_start}:M{investing_row_end})' 
        cashflow_sheet[f'N{cashflow_start_row}'].value = f'=SUM(N{investing_row_start}:N{investing_row_end})' 
        cashflow_sheet[f'O{cashflow_start_row}'].value = f'=SUM(O{investing_row_start}:O{investing_row_end})' 
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(Q{investing_row_start}:Q{investing_row_end})' 

    
        #NET INCREASE Decrease in cash
        cashflow_start_row += 2
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        cashflow_sheet[f'A{cashflow_start_row}'] = 'NET INCREASE (DECREASE IN CASH)'
        cashflow_sheet[f'D{cashflow_start_row}'].value = f'=SUM(D{net_operating_total_row},D{net_investing_total_row})'  
        cashflow_sheet[f'E{cashflow_start_row}'].value = f'=SUM(E{net_operating_total_row},E{net_investing_total_row})'
        cashflow_sheet[f'F{cashflow_start_row}'].value = f'=SUM(F{net_operating_total_row},F{net_investing_total_row})' 
        cashflow_sheet[f'G{cashflow_start_row}'].value = f'=SUM(G{net_operating_total_row},G{net_investing_total_row})' 
        cashflow_sheet[f'H{cashflow_start_row}'].value = f'=SUM(H{net_operating_total_row},H{net_investing_total_row})' 
        cashflow_sheet[f'I{cashflow_start_row}'].value = f'=SUM(I{net_operating_total_row},I{net_investing_total_row})' 
        cashflow_sheet[f'J{cashflow_start_row}'].value = f'=SUM(J{net_operating_total_row},J{net_investing_total_row})' 
        cashflow_sheet[f'K{cashflow_start_row}'].value = f'=SUM(K{net_operating_total_row},K{net_investing_total_row})' 
        cashflow_sheet[f'L{cashflow_start_row}'].value = f'=SUM(L{net_operating_total_row},L{net_investing_total_row})' 
        cashflow_sheet[f'M{cashflow_start_row}'].value = f'=SUM(M{net_operating_total_row},M{net_investing_total_row})' 
        cashflow_sheet[f'N{cashflow_start_row}'].value = f'=SUM(N{net_operating_total_row},N{net_investing_total_row})' 
        cashflow_sheet[f'O{cashflow_start_row}'].value = f'=SUM(O{net_operating_total_row},O{net_investing_total_row})' 
        cashflow_sheet[f'Q{cashflow_start_row}'].value = f'=SUM(Q{net_operating_total_row},Q{net_investing_total_row})' 


        cashflow_start_row  += 2
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        for row in data_balancesheet:
            if row["school"] == school:
                if row['Category'] == 'Assets':
                    if row['Subcategory'] == 'Current Assets':
                        if row['Activity'] == 'Cash':
                            cashflow_sheet[f'A{cashflow_start_row}'] = 'Cash At Beginning of Period'
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['FYE']
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_8']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_10']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'Q{cashflow_start_row}'] = row['FYE']

        cashflow_start_row  += 2
        for col in range(1, 22):
            try:  
                cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                cell.font = fontbold
            except KeyError as e:
                print(f"Error hiding row {col}: {e}")

        for col in range(4, 18):  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.style = currency_style_topdown_border 
        for row in data_balancesheet:
            if row["school"] == school:
                if row['Category'] == 'Assets':
                    if row['Subcategory'] == 'Current Assets':
                        if row['Activity'] == 'Cash':
                            cashflow_sheet[f'A{cashflow_start_row}'] = 'Cash At End of Period'
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_8']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_10']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_6']
                            cashflow_sheet[f'Q{cashflow_start_row}'] = row['last_month_difference']

        while cashflow_start_hiding <= cashflow_start_row:
            for col in range(last_number-3,16):
                col_letter = get_column_letter(col)
             
                cell = cashflow_sheet.cell(row=cashflow_start_hiding, column=col)
                cell.value = None
            cashflow_start_hiding += 1
  


 

    workbook.save(generated_excel_path)

    # Serve the generated Excel file for download
    with open(generated_excel_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(generated_excel_path)}'

    # Remove the generated Excel file (optional)
    os.remove(generated_excel_path)

    return response

def download_csv(request,school):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{school}_GL.csv"'
    print(school)
    cnxn = connect()
    cursor = cnxn.cursor()

    
    current_date = datetime.today().date()   
    current_year = current_date.year
    start_year = current_year
    accper_month_number = int(current_date.strftime("%m"))

    print(accper_month_number)
    if accper_month_number == 1:
        start_year = start_year - 1 

    accper_str = str(accper_month_number).zfill(2)
   
 
    FY_year_1 = start_year
    FY_year_2 = start_year + 1 
    july_date_start  = datetime(FY_year_1, 7, 1).date()
    july_date_end  = datetime(FY_year_2, 6, 30).date()
    september_date_start  = datetime(FY_year_1, 9, 1).date()
    september_date_end  = datetime(FY_year_2, 8, 31).date()

    if school in schoolCategory["ascender"]:
        cursor.execute(
            f"SELECT * FROM [dbo].{db[school]['db']}  as AA where AA.Number != 'BEGBAL';"
        )
    else:
        cursor.execute(f"SELECT * FROM [dbo].{db[school]['db']};")
    rows = cursor.fetchall()
    data3 = []
    
    if school in schoolMonths["julySchool"]:
        current_month = july_date_start
    else:
        current_month = september_date_start
    

    if school in schoolCategory["ascender"]:
        for row in rows:
            expend = float(row[17])
            date = row[11]
            if isinstance(row[11], datetime):
                date = row[11].strftime("%Y-%m-%d")
            acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
            acct_per_month = acct_per_month_string.strftime("%m")
            db_date = row[22].split('-')[0]
            if isinstance(row[11],datetime):
                date_checker = row[11].date()
            else:
                date_checker = datetime.strptime(row[11], "%Y-%m-%d").date()
               

            #convert data
            db_date = int(db_date)
            curr_fy = int(FY_year_1)

                
            if db_date == curr_fy:
                if accper_str != row[12]:
                        
                    row_dict = {
                        "fund": row[0],
                        "func": row[1],
                        "obj": row[2],
                        "sobj": row[3],
                        "org": row[4],
                        "fscl_yr": row[5],
                        "pgm": row[6],
                        "edSpan": row[7],
                        "projDtl": row[8],
                        "AcctDescr": row[9],
                        "Number": row[10],
                        "Date": date,
                        "AcctPer": row[12],
                        "Est": row[13],
                        "Real": row[14],
                        "Appr": row[15],
                        "Encum": row[16],
                        "Expend": expend,
                        "Bal": row[18],
                        "WorkDescr": row[19],
                        "Type": row[20],
                        # "Contr": row[21],
                    }
                    data3.append(row_dict)
            
    
    else:        
        for row in rows:
            amount = float(row[19])
            date = row[9]
            
            if isinstance(row[9], datetime):
                date = row[9].strftime("%Y-%m-%d")
            acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
            acct_per_month = acct_per_month_string.strftime("%m")
            if isinstance(row[9], (datetime, datetime.date)):
                date_checker = row[9].date()
            else:
                date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()
            if school in schoolMonths["julySchool"]:
            
                if date_checker >= july_date_start and date_checker <= july_date_end:
                    if accper_str != row[10]:
                        row_dict = {
                            "fund": row[0],
                            "T":row[1],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Pl":row[7],
                            "LOC":row[8],         
                            "Date": date,
                            "AcctPer":row[10],
                            "Source":row[11],
                            "Subsource":row[12],
                            "Batch":row[13],
                            "Vendor":row[14],
                            "TransactionDescr":row[15],
                            "InvoiceDate":row[16],
                            "CheckNumber":row[17],
                            "CheckDate":row[18],
                            "Amount": amount,
                            "Budget":row[20],
                            "BegBal":row[21],
                        }
                        data3.append(row_dict)
            else:
                if date_checker >= september_date_start and date_checker <= september_date_end:
                    if accper_str != row[10]:
                        row_dict = {
                            "fund": row[0],
                            "T":row[1],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Pl":row[7],
                            "LOC":row[8],         
                            "Date": date,
                            "AcctPer":row[10],
                            "Source":row[11],
                            "Subsource":row[12],
                            "Batch":row[13],
                            "Vendor":row[14],
                            "TransactionDescr":row[15],
                            "InvoiceDate":row[16],
                            "CheckNumber":row[17],
                            "CheckDate":row[18],
                            "Amount": amount,
                            "Budget":row[20],
                            "BegBal":row[21],
                        }
                        data3.append(row_dict)


    csv_writer = csv.writer(response)

    if school in schoolCategory["ascender"]:
        csv_writer.writerow(['fund', 'func', 'obj' , 'sobj', 'org','fscl_yr', 'pgm', 'edSpan', 'projDtl', 'AcctDescr', 'Number', 'Date', 'AcctPer','Est', 'Real' , 'Appr', 'Encum', 'Expend', 'Bal','WorkDescr', 'Type'])  
        for row in data3:
            csv_writer.writerow([row['fund'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['pgm'],row['edSpan'],row['projDtl'],row['AcctDescr'],row['Number'],row['Date'], row['AcctPer'], row['Est'],row['Real'],row['Appr'],row['Encum'],row['Expend'],row['Bal'],row['WorkDescr'],row['Type']])        
    else:
        csv_writer.writerow(['fund', 'T','func', 'obj' , 'sobj', 'org','fscl_yr', 'Pl','LOC','Date', 'AcctPer','Source','Subsource','Batch','Vendor','TransactionDescr','InvoiceDate','CheckNumber','CheckDate', 'Amount', 'Budget','BegBal'])  
        for row in data3:
            csv_writer.writerow([row['fund'],row['T'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['Pl'],row['LOC'],row['Date'],row['AcctPer'],row['Source'],row['Subsource'],row['Batch'],row['Vendor'],row['TransactionDescr'],row['InvoiceDate'],row['CheckNumber'],row['CheckDate'],row['Amount'],row['Budget'],row['BegBal']])
    

    print("done")
    return response




# def upload_drive(request):
#     school = "advantage"
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = f'attachment; filename="{test}_GL.csv"'
    
#     cnxn = connect()
#     cursor = cnxn.cursor()

#     url = f"https://graph.microsoft.com/v1.0/me/drive/items/{onedrive_folder_id}:/{file_path.split('/')[-1]}:/content"
#     headers = {
#         'Authorization': f'Bearer {access_token}',
#         'Content-Type': 'text/csv', 
#     }

#     current_date = datetime.today().date()   
#     current_year = current_date.year
#     start_year = current_year
#     accper_month_number = int(current_date.strftime("%m"))

#     accper_str = str(accper_month_number).zfill(2)
#     print(accper_str)

#     FY_year_1 = start_year
#     FY_year_2 = start_year + 1 
#     july_date_start  = datetime(FY_year_1, 7, 1).date()
#     july_date_end  = datetime(FY_year_2, 6, 30).date()
#     september_date_start  = datetime(FY_year_1, 9, 1).date()
#     september_date_end  = datetime(FY_year_2, 8, 31).date()

#     if school in schoolCategory["ascender"]:
#         cursor.execute(
#             f"SELECT * FROM [dbo].{db[school]['db']}  as AA where AA.Number != 'BEGBAL';"
#         )
#     else:
#         cursor.execute(f"SELECT * FROM [dbo].{db[school]['db']};")
#     rows = cursor.fetchall()
#     data3 = []
    
#     if school in schoolMonths["julySchool"]:
#         current_month = july_date_start
#     else:
#         current_month = september_date_start
    

#     if school in schoolCategory["ascender"]:
#         for row in rows:
#             expend = float(row[17])
#             date = row[11]
#             if isinstance(row[11], datetime):
#                 date = row[11].strftime("%Y-%m-%d")
#             acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
#             acct_per_month = acct_per_month_string.strftime("%m")
#             db_date = row[22].split('-')[0]
#             if isinstance(row[11],datetime):
#                 date_checker = row[11].date()
#             else:
#                 date_checker = datetime.strptime(row[11], "%Y-%m-%d").date()
               

#             #convert data
#             db_date = int(db_date)
#             curr_fy = int(FY_year_1)

                
#             if db_date == curr_fy:
#                 if accper_str != row[12]:
                        
#                     row_dict = {
#                         "fund": row[0],
#                         "func": row[1],
#                         "obj": row[2],
#                         "sobj": row[3],
#                         "org": row[4],
#                         "fscl_yr": row[5],
#                         "pgm": row[6],
#                         "edSpan": row[7],
#                         "projDtl": row[8],
#                         "AcctDescr": row[9],
#                         "Number": row[10],
#                         "Date": date,
#                         "AcctPer": row[12],
#                         "Est": row[13],
#                         "Real": row[14],
#                         "Appr": row[15],
#                         "Encum": row[16],
#                         "Expend": expend,
#                         "Bal": row[18],
#                         "WorkDescr": row[19],
#                         "Type": row[20],
#                         # "Contr": row[21],
#                     }
#                     data3.append(row_dict)
            
    
#     else:        
#         for row in rows:
#             amount = float(row[19])
#             date = row[9]
            
#             if isinstance(row[9], datetime):
#                 date = row[9].strftime("%Y-%m-%d")
#             acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
#             acct_per_month = acct_per_month_string.strftime("%m")
#             if isinstance(row[9], (datetime, datetime.date)):
#                 date_checker = row[9].date()
#             else:
#                 date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()
#             if school in schoolMonths["julySchool"]:
            
#                 if date_checker >= july_date_start and date_checker <= july_date_end:
#                     if accper_str != row[10]:
#                         row_dict = {
#                             "fund": row[0],
#                             "T":row[1],
#                             "func": row[2],
#                             "obj": row[3],
#                             "sobj": row[4],
#                             "org": row[5],
#                             "fscl_yr": row[6],
#                             "Pl":row[7],
#                             "LOC":row[8],         
#                             "Date": date,
#                             "AcctPer":row[10],
#                             "Source":row[11],
#                             "Subsource":row[12],
#                             "Batch":row[13],
#                             "Vendor":row[14],
#                             "TransactionDescr":row[15],
#                             "InvoiceDate":row[16],
#                             "CheckNumber":row[17],
#                             "CheckDate":row[18],
#                             "Amount": amount,
#                             "Budget":row[20],
#                         }
#                         data3.append(row_dict)
#             else:
#                 if date_checker >= september_date_start and date_checker <= september_date_end:
#                     if accper_str != row[10]:
#                         row_dict = {
#                             "fund": row[0],
#                             "T":row[1],
#                             "func": row[2],
#                             "obj": row[3],
#                             "sobj": row[4],
#                             "org": row[5],
#                             "fscl_yr": row[6],
#                             "Pl":row[7],
#                             "LOC":row[8],         
#                             "Date": date,
#                             "AcctPer":row[10],
#                             "Source":row[11],
#                             "Subsource":row[12],
#                             "Batch":row[13],
#                             "Vendor":row[14],
#                             "TransactionDescr":row[15],
#                             "InvoiceDate":row[16],
#                             "CheckNumber":row[17],
#                             "CheckDate":row[18],
#                             "Amount": amount,
#                             "Budget":row[20],
#                         }
#                         data3.append(row_dict)


#     csv_writer = csv.writer(response)

#     if school in schoolCategory["ascender"]:
#         csv_writer.writerow(['fund', 'func', 'obj' , 'sobj', 'org','fscl_yr', 'pgm', 'edSpan', 'projDtl', 'AcctDescr', 'Number', 'Date', 'AcctPer','Est', 'Real' , 'Appr', 'Encum', 'Expend', 'Bal','WorkDescr', 'Type'])  
#         for row in data3:
#             csv_writer.writerow([row['fund'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['pgm'],row['edSpan'],row['projDtl'],row['AcctDescr'],row['Number'],row['Date'], row['AcctPer'], row['Est'],row['Real'],row['Appr'],row['Encum'],row['Expend'],row['Bal'],row['WorkDescr'],row['Type']])        
#     else:
#         csv_writer.writerow(['fund', 'T','func', 'obj' , 'sobj', 'org','fscl_yr', 'Pl','LOC','Date', 'AcctPer','Source','Subsource','Batch','Vendor','TransactionDescr','InvoiceDate','CheckNumber','CheckDate', 'Amount', 'Budget'])  
#         for row in data3:
#             csv_writer.writerow([row['fund'],row['T'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['Pl'],row['LOC'],row['Date'],row['AcctPer'],row['Source'],row['Subsource'],row['Batch'],row['Vendor'],row['TransactionDescr'],row['InvoiceDate'],row['CheckNumber'],row['CheckDate'],row['Amount'],row['Budget']])
    
    
#     print("done")
#     return response


def mockup(request):
    return render(request, "mockup/profit-loss.html")




def upload_data(request,school):



    storage_connection_string = "DefaultEndpointsProtocol=https;AccountName=blogdataprocessing;AccountKey=qI+FDF3NPvo6SkYUpr00yw4MiQB0lofHF+lmnZ+8S686FPXBUTMJZUo31N3KoNefctV/rfR0dTFe+AStBaDTpA==;EndpointSuffix=core.windows.net"
    blob_service_client = BlobServiceClient.from_connection_string(storage_connection_string)

    blob = 'blob-'

    cont_id = blob + school
    container_id = 'clients'
    print(container_id)
    container_client = blob_service_client.get_container_client(container_id)
 

    if request.method == "POST" and request.FILES.getlist('upload'):
        files = request.FILES.getlist('upload')
        ponumber = request.POST.get('ponumber')
        username = request.session.get('username')
        status = "Pending"
        logs = ""
        client = ""
        date = datetime.now()
    
        print(ponumber)

        blobs = [blob.name for blob in container_client.list_blobs()]

        
        for file in files:
            # Get the file name and create a BlobClient
            file_name = file.name
            file_name_hash = hashlib.sha1(file_name.encode()).hexdigest()
            
            
            # if file_name_hash in blobs:
            #     error_message = 'File with the same name already exists.'
            #     return render(request, "temps/data-processing.html", {'error_message': error_message})
            
            blob_client = container_client.get_blob_client(file_name_hash)

            # Upload the file to Azure Storage
            blob_client.upload_blob(file.read(), overwrite=True)
            blob_url = f"{container_id}/{file_name_hash}"
            print(blob_url)

            cnxn = connect()
            cursor = cnxn.cursor()
            insert_query = f"INSERT INTO [dbo].[InvoiceSubmission] (PO_Number, blobPath,client, [user], status, logs,fileName,date)  VALUES (?, ?, ?, ?, ?, ?,?, ?)"
            cursor.execute(insert_query, (ponumber,blob_url,client,username,status,logs,file_name,date))
            cnxn.commit()
            cursor.close()
            cnxn.close()

    return redirect(reverse('data_processing', args=[school]))

def download_file(request, name, school):
    storage_connection_string = "DefaultEndpointsProtocol=https;AccountName=blogdataprocessing;AccountKey=qI+FDF3NPvo6SkYUpr00yw4MiQB0lofHF+lmnZ+8S686FPXBUTMJZUo31N3KoNefctV/rfR0dTFe+AStBaDTpA==;EndpointSuffix=core.windows.net"

    # Create a BlobServiceClient
    blob_service_client = BlobServiceClient.from_connection_string(storage_connection_string)

    
    cont_id = f"blob-{school}" 
    container_id = cont_id

    container_client = blob_service_client.get_container_client(container_id)

    # Get the blob (file) from the container
    name_split = name.split("/")

    file_name = name_split[-1]
    blob_client = container_client.get_blob_client(file_name)

    # Download the blob content
    blob_data = blob_client.download_blob()

    # Create a response with the blob content
    response = HttpResponse(blob_data.readall())
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    return response

